"""
perishable_vrp.py
=================
Multi-Period Collaborative Pickup Open VRP for Perishable Agricultural Products
under Scenario-Based Deterioration Uncertainty

Solvers supported (choose via SOLVER variable at the top):
  - "cplex"   : IBM CPLEX        (pip install cplex  / docplex)
  - "gurobi"  : Gurobi           (pip install gurobipy)
  - "pulp"    : CBC via PuLP     (pip install pulp)            ← free, no licence needed
  - "glpk"    : GLPK via PuLP   (pip install pulp + glpk)     ← free
  - "highs"   : HiGHS via PuLP  (pip install pulp highspy)    ← free, fast

Author: generated from OPL model
"""

import math, random, time, warnings
warnings.filterwarnings("ignore")

# =============================================================================
# 0.  SOLVER SELECTION  ← change this line to switch solver
# =============================================================================
SOLVER = "cplex"          # options: "cplex" | "gurobi" | "pulp" | "glpk" | "highs"

# Time limit (seconds) and MIP gap — respected by all solvers
#TIME_LIMIT = 300         # seconds (5 minutes) — CPLEX stops and returns best solution found
#MIP_GAP    = 0.01        # 1 % optimality gap

# =============================================================================
# 1.  INSTANCE DATA  (edit this section or replace with file I/O)
# =============================================================================
random.seed(42)

# --- Dimensions --------------------------------------------------------------
nF = 5    # number of farmers      (keep small for first test; scale up later)
nK = 2    # number of vehicles
nT = 6    # number of periods
nS = 3    # number of scenarios
nO = nK   # one distinct origin per vehicle

# --- Index sets --------------------------------------------------------------
N   = list(range(1, nF + 1))           # farmers          1 .. nF
K   = list(range(1, nK + 1))           # vehicles         1 .. nK
T   = list(range(1, nT + 1))           # periods          1 .. nT
T0  = list(range(0, nT + 1))           # cohort indices   0 .. nT
S   = list(range(1, nS + 1))           # scenarios        1 .. nS
O   = list(range(nF+1, nF+nO+1))       # origin nodes     nF+1 .. nF+nO
V   = list(range(0, nF+nO+1))          # all nodes        0 (hub) .. nF+nO
HUB = 0                                 # hub node index

# --- Scenario parameters -----------------------------------------------------
rho = {1: 0.30, 2: 0.50, 3: 0.20}     # scenario probabilities
d   = {1: 0.05, 2: 0.10, 3: 0.20}     # deterioration rates

# --- Farmer parameters -------------------------------------------------------
r_price = {i: round(random.uniform(2.50, 3.50), 2) for i in N}  # EUR/crate, fresh unit price
alpha   = {i: round(random.uniform(4.00, 7.00), 2) for i in N}  # EUR/visit service cost
a0      = {i: random.randint(0, 49) for i in N}            # initial inventory

g = {(i, t): max(0, int(random.randint(50, 149) * (1 + 0.1*random.gauss(0, 1))))
     for i in N for t in T}            # harvest quantity

# --- Vehicle parameters ------------------------------------------------------
q        = {k: 245 for k in K}        # crates capacity
originOf = {k: nF + k for k in K}     # origin node of vehicle k

# --- Coordinates & travel costs ----------------------------------------------
# Node 0 = hub (Izmir), 1..nF = farmers (Aydin-Manisa), nF+1..nF+nO = origins
coords = {0: (38.40, 27.14)} #hub location
for i in N:
    coords[i] = (37.50 + random.random(), 27.80 + random.random()*0.6)
    # AFTER — range is explicitly defined
    #coords[i] = (random.uniform(37.50, 38.50),   # latitude
                  #random.uniform(27.80, 28.40))    # longitude
for o in O:
    b = random.choice(N) #pick a random farmer and place the origin near that farmer
    coords[o] = (coords[b][0] + random.gauss(0,0.05),
                 coords[b][1] + random.gauss(0,0.05)) #add +-5 km deviation to that farmer's coordinates

def dist_km(u, v):
    if u == v: return 0.0
    dlat = (coords[v][0]-coords[u][0]) * 111.0 #converts latitude difference to km
    dlon = (coords[v][1]-coords[u][1]) * 111.0 * math.cos(math.radians(38.0)) #converts longitude difference to km
    return math.sqrt(dlat**2 + dlon**2)

cost_per_km = 12/100 * 1.29   # EUR/km  (12 L/100km × 1.29 EUR/L)
c = {(u, v, k): round(dist_km(u,v)*cost_per_km, 4)
     for u in V for v in V for k in K}

# --- Precomputed coefficients ------------------------------------------------
def m_val(i, tau, t, s):
    """Selling price of cohort tau collected in period t under scenario s."""
    return round(r_price[i] * (1 - d[s])**(t - tau), 6)  #34

def ub_cohort_val(i, tau, t, s):  #35
    """Upper bound on available inventory for cohort tau."""
    if tau > t: return 0.0   #not yet harvested
    if tau == 0:  #initial stock
        return round(a0[i] * (1 - d[s])**(t - 1), 6)  #initial stock deteriorates t-1 times by period t
    return round(g[i, tau] * (1 - d[s])**(t - tau), 6) #product harvested in period tau deteriorates (t-tau) times by period t

def ub_total_val(i, t, s):  #36
    """Total inventory upper bound for farmer i at period t."""
    return round(sum(ub_cohort_val(i, tau, t, s) for tau in T0 if tau <= t), 6)

m         = {(i,tau,t,s): m_val(i,tau,t,s)
             for i in N for tau in T0 for t in T for s in S if tau <= t} #gelecekteki cohort toplanamaz
ub_cohort = {(i,tau,t,s): ub_cohort_val(i,tau,t,s)
             for i in N for tau in T0 for t in T for s in S if tau <= t}
ub_total  = {(i,t,s): ub_total_val(i,t,s)
             for i in N for t in T for s in S}

# =============================================================================
# 2.  MODEL BUILDER  (solver-agnostic logic, then solver-specific call)
# =============================================================================

def solve_with_cplex():
    """Build and solve with IBM CPLEX (docplex)."""
    from docplex.mp.model import Model

    mdl = Model(name="Perishable_VRP")
    #mdl.parameters.timelimit    = TIME_LIMIT
    #mdl.parameters.mip.tolerances.mipgap = MIP_GAP

    # ── Decision variables ────────────────────────────────────────────────────
    X = {(u,v,k,t): mdl.binary_var(name=f"X_{u}_{v}_{k}_{t}")
         for u in V for v in V for k in K for t in T if u!=v}
    Z     = {(i,k,t): mdl.binary_var(name=f"Z_{i}_{k}_{t}")
             for i in N for k in K for t in T}
    delta = {(i,t): mdl.binary_var(name=f"delta_{i}_{t}")
             for i in N for t in T}
    W     = {(k,t): mdl.binary_var(name=f"W_{k}_{t}")
             for k in K for t in T}
    B   = {(i,tau,t,s): mdl.continuous_var(lb=0, name=f"B_{i}_{tau}_{t}_{s}")
           for i in N for tau in T0 for t in T for s in S if tau<=t}
    P   = {(i,tau,t,s): mdl.continuous_var(lb=0, name=f"P_{i}_{tau}_{t}_{s}")
           for i in N for tau in T0 for t in T for s in S if tau<=t}
    Inv = {(i,tau,t,s): mdl.continuous_var(lb=0, name=f"Inv_{i}_{tau}_{t}_{s}")
           for i in N for tau in T0 for t in T for s in S if tau<=t}
    Q   = {(i,t,s): mdl.continuous_var(lb=0, name=f"Q_{i}_{t}_{s}")
           for i in N for t in T for s in S}
    L   = {(i,k,t,s): mdl.continuous_var(lb=0, name=f"L_{i}_{k}_{t}_{s}")
           for i in N for k in K for t in T for s in S}
    F   = {(u,v,k,t,s): mdl.continuous_var(lb=0, name=f"F_{u}_{v}_{k}_{t}_{s}")
           for u in V for v in V for k in K for t in T for s in S if u!=v}
    M_aux = {(i,k,t): mdl.continuous_var(lb=0, name=f"Maux_{i}_{k}_{t}")
             for i in N for k in K for t in T}

    # ── Objective ─────────────────────────────────────────────────────────────
    revenue = mdl.sum(rho[s]*m[i,tau,t,s]*P[i,tau,t,s]
                      for i in N for t in T for s in S for tau in T0 if tau<=t)
    travel  = mdl.sum(c[u,v,k]*X[u,v,k,t]
                      for u in V for v in V for k in K for t in T if u!=v)
    service = mdl.sum(alpha[i]*Z[i,k,t] for i in N for k in K for t in T)
    mdl.maximize(revenue - travel - service)

    # ── Constraints ───────────────────────────────────────────────────────────

    # 1. Inventory dynamics
    for i in N:
        for s in S:
            mdl.add_constraint(B[i,0,1,s] == a0[i]) #2
    for i in N:
        for t in T:
            for s in S:
                if t > 1:
                    mdl.add_constraint(B[i,0,t,s] == (1-d[s])*Inv[i,0,t-1,s]) #3
    for i in N:
        for t in T:
            for s in S:
                mdl.add_constraint(B[i,t,t,s] == g[i,t])  #4
    for i in N:
        for t in T:
            for s in S:
                for tau in T:
                    if 1 <= tau < t:
                        mdl.add_constraint(B[i,tau,t,s]==(1-d[s])*Inv[i,tau,t-1,s])  #5

    # 2. Visit activation
    for i in N:
        for t in T:
            mdl.add_constraint(delta[i,t] <= mdl.sum(Z[i,k,t] for k in K))  #6
    for i in N:
        for k in K:
            for t in T:
                mdl.add_constraint(delta[i,t] >= Z[i,k,t])  #7

    # 3. Full pickup
    for i in N:
        for t in T:
            for s in S:
                for tau in T0:
                    if tau <= t:
                        ub = ub_cohort[i,tau,t,s]
                        mdl.add_constraint(P[i,tau,t,s] <= B[i,tau,t,s]) #8
                        mdl.add_constraint(P[i,tau,t,s] <= ub*delta[i,t]) #9
                        mdl.add_constraint(P[i,tau,t,s] >= B[i,tau,t,s]-ub*(1-delta[i,t]))  #10

    # 4. Inventory balance
    for i in N:
        for t in T:
            for s in S:
                for tau in T0:
                    if tau <= t:
                        mdl.add_constraint(Inv[i,tau,t,s]==B[i,tau,t,s]-P[i,tau,t,s])  #11

    # 5. Total collected
    for i in N:
        for t in T:
            for s in S:
                mdl.add_constraint(Q[i,t,s]==mdl.sum(P[i,tau,t,s]
                                   for tau in T0 if tau<=t))   #12

    # 6. Linking
    for i in N:
        for t in T:
            for s in S:
                mdl.add_constraint(mdl.sum(L[i,k,t,s] for k in K)==Q[i,t,s])  #13
    for i in N:
        for k in K:
            for t in T:
                for s in S:
                    mdl.add_constraint(L[i,k,t,s] <= q[k]*Z[i,k,t]) #14
                    mdl.add_constraint(L[i,k,t,s] <= ub_total[i,t,s]*Z[i,k,t])  #15

    # 7. Flow
    for u in V:
        for v in V:
            if u==v: continue
            for k in K:
                for t in T:
                    for s in S:
                        mdl.add_constraint(F[u,v,k,t,s]<=q[k]*X[u,v,k,t]) #16
    for k in K:
        ok = originOf[k]
        for t in T:
            for s in S:
                mdl.add_constraint(mdl.sum(F[ok,v,k,t,s] for v in V if v!=ok)==0)  #17
                mdl.add_constraint(mdl.sum(F[u,ok,k,t,s] for u in V if u!=ok)==0)  #18
                mdl.add_constraint(mdl.sum(F[u,HUB,k,t,s] for u in V if u!=HUB)==
                                   mdl.sum(L[i,k,t,s] for i in N))  #20
                mdl.add_constraint(mdl.sum(F[HUB,v,k,t,s] for v in V if v!=HUB)==0)  #21
    for i in N:
        for k in K:
            for t in T:
                for s in S:
                    mdl.add_constraint(
                        mdl.sum(F[i,v,k,t,s] for v in V if v!=i)
                       -mdl.sum(F[u,i,k,t,s] for u in V if u!=i)
                       == L[i,k,t,s])  #19

    # 8. Open routing
    for k in K:
        ok = originOf[k]
        for t in T:
            mdl.add_constraint(mdl.sum(X[ok,v,k,t] for v in V if v!=ok)==W[k,t])  #22
            mdl.add_constraint(mdl.sum(X[u,HUB,k,t] for u in V if u!=HUB)==W[k,t])  #23
            mdl.add_constraint(mdl.sum(X[u,ok,k,t] for u in V if u!=ok)==0)  #24
            mdl.add_constraint(mdl.sum(X[HUB,v,k,t] for v in V if v!=HUB)==0)  #25

    # 9. Routing conservation
    for i in N:
        for k in K:
            for t in T:
                mdl.add_constraint(mdl.sum(X[u,i,k,t] for u in V if u!=i)==Z[i,k,t])  #26
                mdl.add_constraint(mdl.sum(X[i,v,k,t] for v in V if v!=i)==Z[i,k,t])  #27

    # 10. Subtour elimination
    for i in N:
        for j in N:
            if i==j: continue
            for k in K:
                for t in T:
                    mdl.add_constraint(
                        M_aux[i,k,t]-M_aux[j,k,t]+nF*X[i,j,k,t]
                        <= nF-1+nF*(2-Z[i,k,t]-Z[j,k,t]))   #28

    # 11. Tightening bounds
    for i in N:
        for t in T:
            for s in S:
                for tau in T0:
                    if tau <= t:
                        ub = ub_cohort[i,tau,t,s]
                        mdl.add_constraint(B[i,tau,t,s]   <= ub)
                        mdl.add_constraint(P[i,tau,t,s]   <= ub)
                        mdl.add_constraint(Inv[i,tau,t,s] <= ub)
                mdl.add_constraint(Q[i,t,s] <= ub_total[i,t,s])

    t0  = time.time()
    sol = mdl.solve(log_output=True)
    elapsed = time.time() - t0

    status = str(mdl.solve_details.status) if mdl.solve_details else "Unknown"
    obj    = mdl.objective_value if sol else None

    return mdl, status, obj, elapsed, X, Z, delta, W, B, P, Inv, Q, L, F, M_aux


# =============================================================================
# 3.  EXCEL EXPORT
# =============================================================================

def export_to_excel(status, obj, elapsed, X, Z, delta, W, B, P, Inv, Q, L, F, M_aux,
                    filename="VRP_resultss.xlsx"):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    def val(v):
        if hasattr(v, "varValue"):        return v.varValue
        if hasattr(v, "solution_value"):  return v.solution_value
        if hasattr(v, "X"):               return v.X
        return float(v)

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    center      = Alignment(horizontal="center")

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center

    # ── Summary ───────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    style_header(ws, ["Item", "Value"])
    for r, (item, value) in enumerate([
        ("Solver",     SOLVER.upper()),
        ("Status",     status),
        ("Objective (EUR)", round(obj, 4) if obj is not None else "N/A"),
        ("Wall time (s)",   round(elapsed, 2)),
        ("Farmers (nF)",    nF),
        ("Vehicles (nK)",   nK),
        ("Periods (nT)",    nT),
        ("Scenarios (nS)",  nS),
    ], start=2):
        ws.cell(row=r, column=1, value=item)
        ws.cell(row=r, column=2, value=value)

    # ── Parameters ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("Parameters")
    style_header(ws, ["Parameter", "Index", "Value", "Description"])
    r = 2
    for i in N:
        ws.cell(r, 1, "r_price"); ws.cell(r, 2, f"i={i}")
        ws.cell(r, 3, r_price[i]); ws.cell(r, 4, "EUR/crate, fresh unit price"); r += 1
    for i in N:
        ws.cell(r, 1, "alpha"); ws.cell(r, 2, f"i={i}")
        ws.cell(r, 3, alpha[i]); ws.cell(r, 4, "EUR/visit service cost"); r += 1
    for i in N:
        ws.cell(r, 1, "a0"); ws.cell(r, 2, f"i={i}")
        ws.cell(r, 3, a0[i]); ws.cell(r, 4, "initial inventory (crates)"); r += 1
    for (i, t) in sorted(g.keys()):
        ws.cell(r, 1, "g"); ws.cell(r, 2, f"i={i}, t={t}")
        ws.cell(r, 3, g[i, t]); ws.cell(r, 4, "harvest quantity (crates)"); r += 1
    for k in K:
        ws.cell(r, 1, "q"); ws.cell(r, 2, f"k={k}")
        ws.cell(r, 3, q[k]); ws.cell(r, 4, "vehicle capacity (crates)"); r += 1
    for s in S:
        ws.cell(r, 1, "rho"); ws.cell(r, 2, f"s={s}")
        ws.cell(r, 3, rho[s]); ws.cell(r, 4, "scenario probability"); r += 1
    for s in S:
        ws.cell(r, 1, "d"); ws.cell(r, 2, f"s={s}")
        ws.cell(r, 3, d[s]); ws.cell(r, 4, "deterioration rate"); r += 1

    if obj is None:
        wb.save(filename)
        print(f"  Exported (no solution) -> {filename}")
        return

    # ── X: route arcs ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("X_route_arcs")
    style_header(ws, ["u", "v", "k", "t", "X"])
    r = 2
    for (u, v, k, t), var in X.items():
        v_val = round(val(var), 6)
        if v_val > 1e-6:
            ws.cell(r, 1, u); ws.cell(r, 2, v); ws.cell(r, 3, k)
            ws.cell(r, 4, t); ws.cell(r, 5, v_val)
            r += 1

    # ── Z: vehicle-farmer assignment ──────────────────────────────────────────
    ws = wb.create_sheet("Z_vehicle_farmer")
    style_header(ws, ["i (farmer)", "k (vehicle)", "t (period)", "Z"])
    r = 2
    for (i, k, t), var in Z.items():
        ws.cell(r, 1, i); ws.cell(r, 2, k); ws.cell(r, 3, t)
        ws.cell(r, 4, round(val(var), 6)); r += 1

    # ── delta: farmer visit ────────────────────────────────────────────────────
    ws = wb.create_sheet("delta_farmer_visit")
    style_header(ws, ["i (farmer)", "t (period)", "delta"])
    r = 2
    for (i, t), var in delta.items():
        ws.cell(r, 1, i); ws.cell(r, 2, t)
        ws.cell(r, 3, round(val(var), 6)); r += 1

    # ── W: vehicle active ─────────────────────────────────────────────────────
    ws = wb.create_sheet("W_vehicle_active")
    style_header(ws, ["k (vehicle)", "t (period)", "W"])
    r = 2
    for (k, t), var in W.items():
        ws.cell(r, 1, k); ws.cell(r, 2, t)
        ws.cell(r, 3, round(val(var), 6)); r += 1

    # ── B: available inventory ────────────────────────────────────────────────
    ws = wb.create_sheet("B_available_inv")
    style_header(ws, ["i (farmer)", "tau (cohort)", "t (period)", "s (scenario)", "B"])
    r = 2
    for (i, tau, t, s), var in B.items():
        ws.cell(r, 1, i); ws.cell(r, 2, tau); ws.cell(r, 3, t); ws.cell(r, 4, s)
        ws.cell(r, 5, round(val(var), 4)); r += 1

    # ── P: picked up ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("P_pickup")
    style_header(ws, ["i (farmer)", "tau (cohort)", "t (period)", "s (scenario)", "P"])
    r = 2
    for (i, tau, t, s), var in P.items():
        ws.cell(r, 1, i); ws.cell(r, 2, tau); ws.cell(r, 3, t); ws.cell(r, 4, s)
        ws.cell(r, 5, round(val(var), 4)); r += 1

    # ── Inv: leftover inventory ───────────────────────────────────────────────
    ws = wb.create_sheet("Inv_leftover")
    style_header(ws, ["i (farmer)", "tau (cohort)", "t (period)", "s (scenario)", "Inv"])
    r = 2
    for (i, tau, t, s), var in Inv.items():
        ws.cell(r, 1, i); ws.cell(r, 2, tau); ws.cell(r, 3, t); ws.cell(r, 4, s)
        ws.cell(r, 5, round(val(var), 4)); r += 1

    # ── Q: total collected ────────────────────────────────────────────────────
    ws = wb.create_sheet("Q_total_collected")
    style_header(ws, ["i (farmer)", "t (period)", "s (scenario)", "Q"])
    r = 2
    for (i, t, s), var in Q.items():
        ws.cell(r, 1, i); ws.cell(r, 2, t); ws.cell(r, 3, s)
        ws.cell(r, 4, round(val(var), 4)); r += 1

    # ── L: load per vehicle ───────────────────────────────────────────────────
    ws = wb.create_sheet("L_load")
    style_header(ws, ["i (farmer)", "k (vehicle)", "t (period)", "s (scenario)", "L"])
    r = 2
    for (i, k, t, s), var in L.items():
        ws.cell(r, 1, i); ws.cell(r, 2, k); ws.cell(r, 3, t); ws.cell(r, 4, s)
        ws.cell(r, 5, round(val(var), 4)); r += 1

    # ── F: flow (nonzero only) ────────────────────────────────────────────────
    ws = wb.create_sheet("F_flow")
    style_header(ws, ["u", "v", "k", "t", "s", "F"])
    r = 2
    for (u, v, k, t, s), var in F.items():
        f_val = round(val(var), 4)
        if f_val > 1e-6:
            ws.cell(r, 1, u); ws.cell(r, 2, v); ws.cell(r, 3, k)
            ws.cell(r, 4, t); ws.cell(r, 5, s); ws.cell(r, 6, f_val)
            r += 1

    # ── M_aux: MTZ subtour ────────────────────────────────────────────────────
    ws = wb.create_sheet("Maux_MTZ")
    style_header(ws, ["i (farmer)", "k (vehicle)", "t (period)", "Maux"])
    r = 2
    for (i, k, t), var in M_aux.items():
        ws.cell(r, 1, i); ws.cell(r, 2, k); ws.cell(r, 3, t)
        ws.cell(r, 4, round(val(var), 4)); r += 1

    wb.save(filename)
    print(f"\n  Results exported -> {filename}")


# =============================================================================
# 3.  RESULT PRINTER
# =============================================================================

def print_results(status, obj, elapsed, X, Z, delta, W, B, P, Q, L):
    print("\n" + "="*60)
    print(f"  SOLVER   : {SOLVER.upper()}")
    print(f"  STATUS   : {status}")
    print(f"  OBJECTIVE: {round(obj,4) if obj is not None else 'N/A'} EUR")
    print(f"  WALL TIME: {round(elapsed,2)} s")
    print("="*60)

    if obj is None:
        print("  No feasible solution found.")
        return

    def val(v):
        """Extract numeric value regardless of solver variable type."""
        if hasattr(v, "varValue"):   return v.varValue   # PuLP
        if hasattr(v, "solution_value"): return v.solution_value  # CPLEX
        if hasattr(v, "X"):          return v.X           # Gurobi
        return float(v)

    print("\n--- Active routes (X=1) ---")
    for t in T:
        print(f"  Period {t}:")
        for k in K:
            route_arcs = [(u,v) for (u,v,kk,tt) in X
                          if kk==k and tt==t and val(X[u,v,k,t]) > 0.5]
            if route_arcs:
                print(f"    Vehicle {k}: {route_arcs}")

    print("\n--- Farmer visits (delta=1) ---")
    for t in T:
        visited = [i for i in N if val(delta[i,t]) > 0.5]
        print(f"  Period {t}: farmers {visited}")

    print("\n--- Collected quantities Q[i,t,s] ---")
    for t in T:
        for s in S:
            row_q = {i: round(val(Q[i,t,s]),2) for i in N}
            print(f"  Period {t}, Scenario {s}: {row_q}")


# =============================================================================
# 4.  MAIN ENTRY POINT
# =============================================================================

if __name__ == "__main__":
    print(f"\nBuilding model  |  Solver: {SOLVER.upper()}")
    print(f"Instance: {nF} farmers | {nK} vehicles | {nT} periods | {nS} scenarios")
    print(f"Nodes: {len(V)}  (hub=0, farmers=1..{nF}, origins={nF+1}..{nF+nO})\n")

    if SOLVER == "pulp":
        mdl, status, obj, elapsed, X, Z, delta, W, B, P, Inv, Q, L, F = \
            solve_with_pulp("CBC")

    elif SOLVER == "glpk":
        mdl, status, obj, elapsed, X, Z, delta, W, B, P, Inv, Q, L, F = \
            solve_with_pulp("GLPK")

    elif SOLVER == "highs":
        mdl, status, obj, elapsed, X, Z, delta, W, B, P, Inv, Q, L, F = \
            solve_with_pulp("HiGHS")

    elif SOLVER == "gurobi":
        mdl, status, obj, elapsed, X, Z, delta, W, B, P, Inv, Q, L, F = \
            solve_with_gurobi()

    elif SOLVER == "cplex":
        mdl, status, obj, elapsed, X, Z, delta, W, B, P, Inv, Q, L, F, M_aux = \
            solve_with_cplex()

    else:
        raise ValueError(f"Unknown solver '{SOLVER}'. "
                         f"Choose: pulp | glpk | highs | gurobi | cplex")

    print_results(status, obj, elapsed, X, Z, delta, W, B, P, Q, L)
    export_to_excel(status, obj, elapsed, X, Z, delta, W, B, P, Inv, Q, L, F, M_aux)

'''
    def solve_with_gurobi():
    """Build and solve with Gurobi (gurobipy)."""
    import gurobipy as gp
    from gurobipy import GRB

    mdl = gp.Model("Perishable_VRP")
    mdl.setParam("TimeLimit",   TIME_LIMIT)
    mdl.setParam("MIPGap",      MIP_GAP)
    mdl.setParam("OutputFlag",  1)

    # ── Decision variables ────────────────────────────────────────────────────
    X = mdl.addVars([(u,v,k,t) for u in V for v in V for k in K for t in T if u!=v],
                    vtype=GRB.BINARY, name="X")
    Z     = mdl.addVars(N, K, T, vtype=GRB.BINARY,      name="Z")
    delta = mdl.addVars(N, T,    vtype=GRB.BINARY,      name="delta")
    W     = mdl.addVars(K, T,    vtype=GRB.BINARY,      name="W")

    B   = mdl.addVars([(i,tau,t,s) for i in N for tau in T0
                        for t in T for s in S if tau<=t], lb=0, name="B")
    P   = mdl.addVars([(i,tau,t,s) for i in N for tau in T0
                        for t in T for s in S if tau<=t], lb=0, name="P")
    Inv = mdl.addVars([(i,tau,t,s) for i in N for tau in T0
                        for t in T for s in S if tau<=t], lb=0, name="Inv")
    Q   = mdl.addVars(N, T, S, lb=0, name="Q")
    L   = mdl.addVars(N, K, T, S, lb=0, name="L")
    F   = mdl.addVars([(u,v,k,t,s) for u in V for v in V for k in K
                        for t in T for s in S if u!=v], lb=0, name="F")
    M_aux = mdl.addVars(N, K, T, lb=0, name="Maux")

    # ── Objective ─────────────────────────────────────────────────────────────
    revenue = gp.quicksum(rho[s]*m[i,tau,t,s]*P[i,tau,t,s]
                          for i in N for t in T for s in S
                          for tau in T0 if tau<=t)
    travel  = gp.quicksum(c[u,v,k]*X[u,v,k,t]
                          for u in V for v in V for k in K for t in T if u!=v)
    service = gp.quicksum(alpha[i]*Z[i,k,t]
                          for i in N for k in K for t in T)
    mdl.setObjective(revenue - travel - service, GRB.MAXIMIZE)

    # ── Constraints ───────────────────────────────────────────────────────────

    # 1. Inventory dynamics
    for i in N:
        for s in S:
            mdl.addConstr(B[i,0,1,s] == a0[i],                   f"B_init_{i}_{s}")
    for i in N:
        for t in T:
            for s in S:
                if t > 1:
                    mdl.addConstr(B[i,0,t,s] == (1-d[s])*Inv[i,0,t-1,s],
                                  f"B_c0_{i}_{t}_{s}")
    for i in N:
        for t in T:
            for s in S:
                mdl.addConstr(B[i,t,t,s] == g[i,t],              f"B_harv_{i}_{t}_{s}")
    for i in N:
        for t in T:
            for s in S:
                for tau in T:
                    if 1 <= tau < t:
                        mdl.addConstr(B[i,tau,t,s]==(1-d[s])*Inv[i,tau,t-1,s],
                                      f"B_old_{i}_{tau}_{t}_{s}")

    # 2. Visit activation
    for i in N:
        for t in T:
            mdl.addConstr(delta[i,t] <= gp.quicksum(Z[i,k,t] for k in K),
                          f"dlt_ub_{i}_{t}")
    for i in N:
        for k in K:
            for t in T:
                mdl.addConstr(delta[i,t] >= Z[i,k,t],            f"dlt_lb_{i}_{k}_{t}")

    # 3. Full pickup
    for i in N:
        for t in T:
            for s in S:
                for tau in T0:
                    if tau <= t:
                        ub = ub_cohort[i,tau,t,s]
                        mdl.addConstr(P[i,tau,t,s] <= B[i,tau,t,s],
                                      f"pu_B_{i}_{tau}_{t}_{s}")
                        mdl.addConstr(P[i,tau,t,s] <= ub*delta[i,t],
                                      f"pu_D_{i}_{tau}_{t}_{s}")
                        mdl.addConstr(P[i,tau,t,s] >= B[i,tau,t,s]-ub*(1-delta[i,t]),
                                      f"pu_lb_{i}_{tau}_{t}_{s}")

    # 4. Inventory balance
    for i in N:
        for t in T:
            for s in S:
                for tau in T0:
                    if tau <= t:
                        mdl.addConstr(Inv[i,tau,t,s]==B[i,tau,t,s]-P[i,tau,t,s],
                                      f"invbal_{i}_{tau}_{t}_{s}")

    # 5. Total collected
    for i in N:
        for t in T:
            for s in S:
                mdl.addConstr(Q[i,t,s]==gp.quicksum(P[i,tau,t,s]
                              for tau in T0 if tau<=t),           f"Q_{i}_{t}_{s}")

    # 6. Linking
    for i in N:
        for t in T:
            for s in S:
                mdl.addConstr(gp.quicksum(L[i,k,t,s] for k in K)==Q[i,t,s],
                              f"lnk_spl_{i}_{t}_{s}")
    for i in N:
        for k in K:
            for t in T:
                for s in S:
                    mdl.addConstr(L[i,k,t,s] <= q[k]*Z[i,k,t],  f"lnk_cap_{i}_{k}_{t}_{s}")
                    mdl.addConstr(L[i,k,t,s] <= ub_total[i,t,s]*Z[i,k,t],f"lnk_vis_{i}_{k}_{t}_{s}")

    # 7. Flow
    for u in V:
        for v in V:
            if u==v: continue
            for k in K:
                for t in T:
                    for s in S:
                        mdl.addConstr(F[u,v,k,t,s]<=q[k]*X[u,v,k,t],
                                      f"fc_{u}_{v}_{k}_{t}_{s}")
    for k in K:
        ok = originOf[k]
        for t in T:
            for s in S:
                mdl.addConstr(gp.quicksum(F[ok,v,k,t,s] for v in V if v!=ok)==0,
                              f"fo_out_{k}_{t}_{s}")
                mdl.addConstr(gp.quicksum(F[u,ok,k,t,s] for u in V if u!=ok)==0,
                              f"fo_in_{k}_{t}_{s}")
                mdl.addConstr(gp.quicksum(F[u,HUB,k,t,s] for u in V if u!=HUB)==
                              gp.quicksum(L[i,k,t,s] for i in N),
                              f"fh_in_{k}_{t}_{s}")
                mdl.addConstr(gp.quicksum(F[HUB,v,k,t,s] for v in V if v!=HUB)==0,
                              f"fh_out_{k}_{t}_{s}")
    for i in N:
        for k in K:
            for t in T:
                for s in S:
                    mdl.addConstr(
                        gp.quicksum(F[i,v,k,t,s] for v in V if v!=i)
                       -gp.quicksum(F[u,i,k,t,s] for u in V if u!=i)
                       == L[i,k,t,s],                            f"fbal_{i}_{k}_{t}_{s}")

    # 8. Open routing
    for k in K:
        ok = originOf[k]
        for t in T:
            mdl.addConstr(gp.quicksum(X[ok,v,k,t] for v in V if v!=ok)==W[k,t],
                          f"op_dep_{k}_{t}")
            mdl.addConstr(gp.quicksum(X[u,HUB,k,t] for u in V if u!=HUB)==W[k,t],
                          f"op_arr_{k}_{t}")
            mdl.addConstr(gp.quicksum(X[u,ok,k,t] for u in V if u!=ok)==0,
                          f"no_ret_{k}_{t}")
            mdl.addConstr(gp.quicksum(X[HUB,v,k,t] for v in V if v!=HUB)==0,
                          f"no_lhub_{k}_{t}")

    # 9. Routing conservation
    for i in N:
        for k in K:
            for t in T:
                mdl.addConstr(gp.quicksum(X[u,i,k,t] for u in V if u!=i)==Z[i,k,t],
                              f"rin_{i}_{k}_{t}")
                mdl.addConstr(gp.quicksum(X[i,v,k,t] for v in V if v!=i)==Z[i,k,t],
                              f"rout_{i}_{k}_{t}")

    # 10. Subtour elimination
    for i in N:
        for j in N:
            if i==j: continue
            for k in K:
                for t in T:
                    mdl.addConstr(
                        M_aux[i,k,t]-M_aux[j,k,t]+nF*X[i,j,k,t]
                        <= nF-1+nF*(2-Z[i,k,t]-Z[j,k,t]),       f"mtz_{i}_{j}_{k}_{t}")

    # 11. Tightening bounds
    for i in N:
        for t in T:
            for s in S:
                for tau in T0:
                    if tau <= t:
                        ub = ub_cohort[i,tau,t,s]
                        mdl.addConstr(B[i,tau,t,s]   <= ub,      f"tB_{i}_{tau}_{t}_{s}")
                        mdl.addConstr(P[i,tau,t,s]   <= ub,      f"tP_{i}_{tau}_{t}_{s}")
                        mdl.addConstr(Inv[i,tau,t,s] <= ub,      f"tI_{i}_{tau}_{t}_{s}")
                mdl.addConstr(Q[i,t,s] <= ub_total[i,t,s],       f"tQ_{i}_{t}_{s}")

    t0 = time.time()
    mdl.optimize()
    elapsed = time.time() - t0

    status_map = {2:"Optimal", 3:"Infeasible", 5:"Unbounded", 9:"TimeLimit"}
    status = status_map.get(mdl.Status, f"Code_{mdl.Status}")
    obj    = mdl.ObjVal if mdl.SolCount > 0 else None

    return mdl, status, obj, elapsed, X, Z, delta, W, B, P, Inv, Q, L, F
'''