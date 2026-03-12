import random
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =========================
# 0) INPUT: mevcut 30 rota
# =========================
'''
route_to_am_30 = {
    1:2, 2:1, 3:1, 4:2, 5:3, 6:3, 7:2, 8:3, 9:2, 10:1,
    11:2, 12:3, 13:2, 14:1, 15:3, 16:2, 17:2, 18:2, 19:1, 20:3,
    21:2, 22:3, 23:3, 24:2, 25:1, 26:1, 27:3, 28:1, 29:1, 30:1
}

route_to_depots_30 = {
    1:[25,24,14,13],
    2:[16,17,23,20],
    3:[10,22,27,30],
    4:[28,26,19],
    5:[11,12,29],
    6:[8,18,21,15,9],
    7:[19,20,14,7,2],
    8:[17,15,21,25],
    9:[22,8,6,5,1],
    10:[23,13,10,11],
    11:[29,26,24,16,4,3],
    12:[9,18,30,28,27],
    13:[30,23,5,4,2],
    14:[12,15,21,25,28],
    15:[7,9,10,16,22],
    16:[18,19,17,14,12],
    17:[29,24,8,1,3],
    18:[27,20,7,6,4,2],
    19:[7,6,5,3,1],
    20:[20,21,22,24,26,28],
    21:[23,8,16,13,11],
    22:[6,5,4,3,2,1],
    23:[17,30,14,9,10,12],
    24:[18,19,25,26,27,29],
    25:[11,13,15,20,22,26],
    26:[10,16,15,27],
    27:[7,13,14,17,28],
    28:[9,12,11],
    29:[8,19,23,25,24,21],
    30:[18,30,29]
}

ALL_DEPOTS = list(range(1, 31))

# =========================
# 1) yardımcı fonksiyonlar (rota üretimi)
# =========================
def normalize_route(depots):
    """Tekrarları at, sırayı koru."""
    seen = set()
    out = []
    for d in depots:
        if d not in seen:
            out.append(d)
            seen.add(d)
    return out

def mutate_route(depots, hub, min_len=3, max_len=6):
    depots = depots[:]
    op = random.choice(["swap", "add", "remove", "replace"])

    if op == "swap" and len(depots) >= 2:
        i, j = random.sample(range(len(depots)), 2)
        depots[i], depots[j] = depots[j], depots[i]

    elif op == "add" and len(depots) < max_len:
        cand = [d for d in ALL_DEPOTS if d not in depots]
        if cand:
            depots.insert(random.randrange(len(depots)+1), random.choice(cand))

    elif op == "remove" and len(depots) > min_len:
        depots.pop(random.randrange(len(depots)))

    elif op == "replace":
        cand = [d for d in ALL_DEPOTS if d not in depots]
        if cand and depots:
            depots[random.randrange(len(depots))] = random.choice(cand)

    depots = normalize_route(depots)

    if len(depots) < min_len:
        cand = [d for d in ALL_DEPOTS if d not in depots]
        random.shuffle(cand)
        depots += cand[:(min_len - len(depots))]
    if len(depots) > max_len:
        depots = depots[:max_len]

    return depots

def random_route(min_len=3, max_len=6):
    L = random.randint(min_len, max_len)
    return random.sample(ALL_DEPOTS, L)

def merge_routes(a, b, min_len=3, max_len=6):
    merged = normalize_route(a + b)
    if len(merged) < min_len:
        cand = [d for d in ALL_DEPOTS if d not in merged]
        random.shuffle(cand)
        merged += cand[:(min_len - len(merged))]
    if len(merged) > max_len:
        merged = merged[:max_len]
    return merged

def signature(hub, depots):
    """Hub + depo kümesi (sıra önemsiz)"""
    return (hub, tuple(sorted(depots)))

def expand_routes_to_target(route_to_am, route_to_depots,
                             target=600, seed=42,
                             min_len=3, max_len=6,
                             n_mutate=220, n_random=140, n_merge=110):

    random.seed(seed)

    new_am = dict(route_to_am)
    new_dep = {r: normalize_route(ds) for r, ds in route_to_depots.items()}
    used = set(signature(new_am[r], new_dep[r]) for r in new_dep)

    next_id = max(new_dep.keys()) + 1

    # (A) mutation
    base_routes = list(new_dep.keys())
    for _ in range(n_mutate):
        if next_id > target:
            break
        r0 = random.choice(base_routes)
        hub = new_am[r0]
        depots = mutate_route(new_dep[r0], hub, min_len, max_len)
        sig = signature(hub, depots)
        if sig in used:
            continue
        new_am[next_id] = hub
        new_dep[next_id] = depots
        used.add(sig)
        next_id += 1

    # (B) random
    for _ in range(n_random):
        if next_id > target:
            break
        hub = random.choice([1,2,3])
        depots = random_route(min_len, max_len)
        sig = signature(hub, depots)
        if sig in used:
            continue
        new_am[next_id] = hub
        new_dep[next_id] = depots
        used.add(sig)
        next_id += 1

    # (C) merge (same hub)
    by_hub = {1:[], 2:[], 3:[]}
    for r in new_dep:
        by_hub[new_am[r]].append(r)

    for _ in range(n_merge):
        if next_id > target:
            break
        hub = random.choice([1,2,3])
        if len(by_hub[hub]) < 2:
            continue
        r1, r2 = random.sample(by_hub[hub], 2)
        depots = merge_routes(new_dep[r1], new_dep[r2], min_len, max_len)
        sig = signature(hub, depots)
        if sig in used:
            continue
        new_am[next_id] = hub
        new_dep[next_id] = depots
        used.add(sig)
        by_hub[hub].append(next_id)
        next_id += 1

    # Fill remaining up to target
    attempts = 0
    while next_id <= target:
        attempts += 1
        if attempts > 100000:
            print(f"WARNING: Could only generate {next_id - 1} unique routes before exhausting attempts.")
            break
        if random.random() < 0.5:
            r0 = random.choice(list(new_dep.keys()))
            hub = new_am[r0]
            depots = mutate_route(new_dep[r0], hub, min_len, max_len)
        else:
            hub = random.choice([1,2,3])
            depots = random_route(min_len, max_len)

        sig = signature(hub, depots)
        if sig in used:
            continue
        new_am[next_id] = hub
        new_dep[next_id] = depots
        used.add(sig)
        next_id += 1

    return new_am, new_dep

# =========================
# 2) 500 rota üret
# =========================
route_to_am_600, route_to_depots_600 = expand_routes_to_target(
    route_to_am_30, route_to_depots_30,
    target=600, seed=42,
    min_len=3, max_len=6,
    n_mutate=220, n_random=240, n_merge=110
)

print(f"Toplam üretilen rota sayısı: {len(route_to_depots_600)}")

# =========================
# 3) rotalarr.csv formatına çevir
# =========================
RETURN_TO_HUB = False

def build_route_points(hub_id: int, depots: list) -> list:
    pts = [f"Bölge birliği {hub_id}"] + [f"Depo {d}" for d in depots]
    if RETURN_TO_HUB:
        pts.append(f"Bölge birliği {hub_id}")
    return pts

rotalarr_lines = []
for rid in range(1, 601):
    hub_id = route_to_am_600[rid]
    depots = route_to_depots_600[rid]
    points = build_route_points(hub_id, depots)
    line = ",".join([f"Rota {rid}"] + points)
    rotalarr_lines.append(line)

rotalarr_out = "rotalarr_600_generated.csv"
with open(rotalarr_out, "w", encoding="utf-8-sig") as f:
    f.write("\n".join(rotalarr_lines))

print("CSV yazıldı: rotalarr_600_generated.csv")


# =========================
# 4) Excel çıktısı — 4 sayfa
# =========================

# --- Stil sabitleri ---
HEADER_FILL  = PatternFill("solid", start_color="2E4057")   # koyu lacivert
SUBHEAD_FILL = PatternFill("solid", start_color="048A81")   # teal
ALT_FILL     = PatternFill("solid", start_color="EAF4F4")   # açık mint
WHITE_FILL   = PatternFill("solid", start_color="FFFFFF")
HUB_COLORS   = {1: "F4A261", 2: "2A9D8F", 3: "E76F51"}     # hub renk rozeti

thin   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_font(size=11, bold=True, color="FFFFFF"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="1A1A2E"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def style_header_row(ws, row, cols, fill=HEADER_FILL):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = hdr_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def style_data_cell(ws, row, col, value, alt=False, bold=False, align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = ALT_FILL if alt else WHITE_FILL
    cell.font = body_font(bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = BORDER
    return cell

def add_title_banner(ws, text, merge_range):
    ws.merge_cells(merge_range)
    first_cell = ws[merge_range.split(":")[0]]
    first_cell.value = text
    first_cell.font = hdr_font(size=14)
    first_cell.fill = HEADER_FILL
    first_cell.alignment = Alignment(horizontal="center", vertical="center")

wb = Workbook()

# ----------------------------------------------------------
# Sayfa 1 — Ham Veri
# ----------------------------------------------------------
ws1 = wb.active
ws1.title = "Raw Data"
ws1.freeze_panes = "A3"
ws1.row_dimensions[1].height = 30
ws1.row_dimensions[2].height = 20

add_title_banner(ws1, "600 Route Dataset — Raw Data", "A1:I1")

headers1 = ["Route No", "Starting Hub", "No. of Depots",
            "Depot 1", "Depot 2", "Depot 3", "Depot 4", "Depot 5", "Depot 6"]
for col, h in enumerate(headers1, 1):
    ws1.cell(row=2, column=col, value=h)
style_header_row(ws1, 2, len(headers1), fill=SUBHEAD_FILL)

for i, rid in enumerate(range(1, 601)):
    row = i + 3
    alt = (i % 2 == 1)
    hub    = route_to_am_600[rid]
    depots = route_to_depots_600[rid]

    style_data_cell(ws1, row, 1, rid,        alt, bold=True)
    hub_cell = style_data_cell(ws1, row, 2, hub, alt)
    hub_cell.fill = PatternFill("solid", start_color=HUB_COLORS[hub])
    hub_cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    style_data_cell(ws1, row, 3, len(depots), alt)
    for j, d in enumerate(depots):
        style_data_cell(ws1, row, 4 + j, d, alt)
    for j in range(len(depots), 6):
        style_data_cell(ws1, row, 4 + j, "", alt)

for i, w in enumerate([10, 14, 13, 9, 9, 9, 9, 9, 9], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ----------------------------------------------------------
# Sayfa 2 — Python Stili:  route_no: hub  /  route_no: [depots]
# ----------------------------------------------------------
ws2 = wb.create_sheet("Python Style")
ws2.freeze_panes = "A3"
ws2.row_dimensions[1].height = 30
ws2.row_dimensions[2].height = 20

add_title_banner(ws2, "Python Style Notation", "A1:C1")

for col, h in enumerate(["Route No", "route_no: hub", "route_no: [depots]"], 1):
    ws2.cell(row=2, column=col, value=h)
style_header_row(ws2, 2, 3, fill=SUBHEAD_FILL)

for i, rid in enumerate(range(1, 601)):
    row = i + 3
    alt    = (i % 2 == 1)
    hub    = route_to_am_600[rid]
    depots = route_to_depots_600[rid]
    style_data_cell(ws2, row, 1, rid,                    alt, bold=True)
    style_data_cell(ws2, row, 2, f"{rid}: {hub}",        alt, align="left")
    style_data_cell(ws2, row, 3, f"{rid}: {depots}",     alt, align="left")

ws2.column_dimensions["A"].width = 10
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 55

# ----------------------------------------------------------
# Sayfa 3 — OPL Stili:  <route_no, hub, {depots}>
# ----------------------------------------------------------
ws3 = wb.create_sheet("OPL Style")
ws3.freeze_panes = "A3"
ws3.row_dimensions[1].height = 30
ws3.row_dimensions[2].height = 20

add_title_banner(ws3, "OPL Style Notation", "A1:B1")

for col, h in enumerate(["Route No", "<route_no, hub, {depots}>"], 1):
    ws3.cell(row=2, column=col, value=h)
style_header_row(ws3, 2, 2, fill=SUBHEAD_FILL)

for i, rid in enumerate(range(1, 601)):
    row = i + 3
    alt    = (i % 2 == 1)
    hub    = route_to_am_600[rid]
    depots = route_to_depots_600[rid]
    depot_set = "{" + ",".join(str(d) for d in depots) + "}"
    opl_str   = f"<{rid},{hub},{depot_set}>"
    style_data_cell(ws3, row, 1, rid,     alt, bold=True)
    style_data_cell(ws3, row, 2, opl_str, alt, align="left")

ws3.column_dimensions["A"].width = 10
ws3.column_dimensions["B"].width = 55

# ----------------------------------------------------------
# Sayfa 4 — Birleşik Görünüm (tüm notasyonlar yan yana)
# ----------------------------------------------------------
ws4 = wb.create_sheet("Combined View")
ws4.freeze_panes = "A3"
ws4.row_dimensions[1].height = 30
ws4.row_dimensions[2].height = 20

add_title_banner(ws4, "All Notation Styles — Combined View", "A1:D1")

for col, h in enumerate(["Route No", "route_no: hub",
                          "route_no: [depots]", "<route_no, hub, {depots}>"], 1):
    ws4.cell(row=2, column=col, value=h)
style_header_row(ws4, 2, 4, fill=SUBHEAD_FILL)

for i, rid in enumerate(range(1, 601)):
    row = i + 3
    alt    = (i % 2 == 1)
    hub    = route_to_am_600[rid]
    depots = route_to_depots_600[rid]
    depot_set = "{" + ",".join(str(d) for d in depots) + "}"
    opl_str   = f"<{rid},{hub},{depot_set}>"
    style_data_cell(ws4, row, 1, rid,                alt, bold=True)
    style_data_cell(ws4, row, 2, f"{rid}: {hub}",   alt, align="left")
    style_data_cell(ws4, row, 3, f"{rid}: {depots}", alt, align="left")
    style_data_cell(ws4, row, 4, opl_str,            alt, align="left")

ws4.column_dimensions["A"].width = 10
ws4.column_dimensions["B"].width = 18
ws4.column_dimensions["C"].width = 50
ws4.column_dimensions["D"].width = 55

# ----------------------------------------------------------
# Kaydet
# ----------------------------------------------------------
excel_out = "routes_600.xlsx"
wb.save(excel_out)
print(f"Excel yazıldı: {excel_out}")
'''


#r'''
# =========================
# 3) Noktalar.csv oku (birebir) + geodesic
# =========================
from geopy.distance import geodesic
import pandas as pd
import os

# ---- DOSYA YOLLARI ----
noktalar_file = r"C:\Users\gizem\OneDrive\Belgeler\GitHub\gizem_tez\Makale-21-01-26\Noktalar.csv"
rotalar_file  = r"C:\Users\gizem\OneDrive\Belgeler\GitHub\gizem_tez\Makale-21-01-26\rotalarr600.csv"

def clean_name(x):
    return str(x).replace("\u00a0", " ").strip()

# =========================
# 1) NOKTALAR VERİSİNİ HAZIRLA
# =========================
df_noktalar = pd.read_csv(noktalar_file, encoding="utf-8-sig")
df_noktalar["Ad"] = df_noktalar["Ad"].apply(clean_name)

# Koordinat temizliği
df_noktalar["Enlem"]  = pd.to_numeric(df_noktalar["Enlem"].astype(str).str.replace(",", "."), errors="coerce")
df_noktalar["Boylam"] = pd.to_numeric(df_noktalar["Boylam"].astype(str).str.replace(",", "."), errors="coerce")

if df_noktalar[["Enlem", "Boylam"]].isna().any().any():
    print("Uyarı: Bazı koordinatlar sayıya dönüştürülemedi!")

noktalar_dict = df_noktalar.set_index("Ad")[["Enlem", "Boylam"]].to_dict(orient="index")

# =========================
# 2) ROTALAR VERİSİNİ OKU
# =========================
with open(rotalar_file, "r", encoding="utf-8-sig") as f:
    max_columns = max(len(line.strip().split(",")) for line in f)

column_names = ["Route"] + [f"Point_{i}" for i in range(1, max_columns)]
df_rotalar = pd.read_csv(rotalar_file, encoding="utf-8-sig", names=column_names, header=None, engine="python")

# Veri temizleme
for c in df_rotalar.columns:
    df_rotalar[c] = df_rotalar[c].apply(clean_name)
    df_rotalar.loc[df_rotalar[c].isin(["", "nan", "None", "None"]), c] = None

# =========================
# 3) HESAPLAMA FONKSİYONLARI
# =========================
fuel_price_per_litre = 1.29
capacity_by_num_depots = {3: 750, 4: 900, 5: 1200, 6: 1400}
# Ziyaret edilen depo sayısına göre 100km'deki yakıt tüketimi (Litre)
fuel_consumption_map = {
    3: 25,
    4: 30,
    5: 38,
    6: 38
}

def calculate_route_distance(route_points):
    total = 0.0
    for i in range(len(route_points) - 1):
        p1, p2 = route_points[i], route_points[i+1]
        a, b = noktalar_dict.get(p1), noktalar_dict.get(p2)
        if not a or not b:
            missing = p1 if not a else p2
            raise KeyError(f"Nokta bulunamadı: {missing}")
        total += geodesic((a["Enlem"], a["Boylam"]), (b["Enlem"], b["Boylam"])).km
    return total

def route_cost(distance_km, num_depots):
    # Tablodan o depo sayısına denk gelen tüketim oranını al (Bulamazsa varsayılan 30 al)
    consumption = fuel_consumption_map.get(num_depots, 40)
    
    # Hesaplama: (Mesafe * (Tüketim / 100)) * Yakıt Fiyatı
    return (distance_km * (consumption / 100.0)) * fuel_price_per_litre

def get_capacity(route_points):
    # 'Depo' veya 'D' ile başlayan noktaları say (Hub hariç)
    # Senin formatına göre 'Depo' kelimesini aratıyoruz
    n_depots = sum(1 for p in route_points if "DEPO" in str(p).upper())
    return capacity_by_num_depots.get(n_depots, None)

# =========================
# 4) ANA DÖNGÜ VE RAPORLAMA
# =========================
rows_out = []

for _, row in df_rotalar.iterrows():
    route_name = row["Route"]
    # Satırdaki boş olmayan tüm noktaları listeye al (Rota adı hariç)
    all_points = row.drop(labels=["Route"]).dropna().tolist()
    
    try:
        dist_km = calculate_route_distance(all_points)
        
        # DÜZELTME: Maliyet hesabı için önce depo sayısını buluyoruz
        depots_only_count = sum(1 for p in all_points if "DEPO" in str(p).upper())
        cost_eur = route_cost(dist_km, depots_only_count) # Fonksiyona ikinci parametre eklendi
        
        cap = get_capacity(all_points)
        err = ""
    except Exception as e:
        dist_km, cost_eur, cap = None, None, None
        err = str(e)

    # Hub ve Depo bilgisini görselleştirmek için ayıkla
    hub_point = all_points[0] if all_points else "Bilinmiyor"
    depots_only = [p for p in all_points if "DEPO" in str(p).upper()]

    rows_out.append({
        "route_name": route_name,
        "hub_point": hub_point,
        "depots": " - ".join(depots_only),
        "num_depots": len(depots_only),
        "distance_km": dist_km,
        "route_cost_eur": cost_eur,
        "route_capacity": cap,
        "error": err
    })

df_out = pd.DataFrame(rows_out)

# =========================
# 5) ÇIKTI
# =========================
output_excel = "routes_analysis_results600.xlsx"
with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
    df_out.to_excel(writer, sheet_name="Analiz Sonuclari", index=False)
    df_rotalar.to_excel(writer, sheet_name="Okunan Rotalar", index=False)

print(f"✅ Analiz tamamlandı ve '{output_excel}' dosyasına kaydedildi.")
print(df_out[["route_name", "distance_km", "route_cost_eur"]].head())
#'''
