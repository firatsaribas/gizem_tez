"""
perishable_vrp_kg.py
====================
Multi-Period Collaborative Pickup Open VRP for Perishable Agricultural Products
under Scenario-Based Deterioration Uncertainty

Unit: KILOGRAMS (converted from crate-based model)
Conversion: 1 crate = 20 kg (standard Guloglu 530x365x315 mm tomato crate)
Vehicle capacity: 3,500 kg (weight-limited pickup truck payload)

Solver: IBM CPLEX (docplex)
"""

import math, time, warnings, os
import openpyxl

warnings.filterwarnings("ignore")

SOLVER = "cplex"
PARAMS_FILE = "parameters_input.xlsx"  # <-- change path if needed


# =============================================================================
# 1.  READ PARAMETERS FROM EXCEL
# =============================================================================

def load_parameters(filepath):
    """
    Read all instance parameters from the structured Excel workbook.
    Returns a dict with keys: nF, nK, nT, nS, rho, d, q, r_price,
    alpha, a0, g, coords.
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(
            f"Parameter file '{filepath}' not found.\n"
            f"Please create it using the provided template."
        )

    wb = openpyxl.load_workbook(filepath, data_only=True)
    params = {}

    # ── Sheet 1: Dimensions ──────────────────────────────────────────────────
    ws = wb["Dimensions"]
    dim = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1] is not None:
            dim[str(row[0]).strip()] = int(row[1])

    params["nF"] = dim["nF"]
    params["nK"] = dim["nK"]
    params["nT"] = dim["nT"]
    params["nS"] = dim["nS"]

    N = list(range(1, params["nF"] + 1))
    K = list(range(1, params["nK"] + 1))
    T = list(range(1, params["nT"] + 1))
    S = list(range(1, params["nS"] + 1))

    # ── Sheet 2: Scenarios ───────────────────────────────────────────────────
    ws = wb["Scenarios"]
    d = {}
    rho = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            s = int(row[0])
            d[s] = float(row[1])
            rho[s] = float(row[2])

    # Warn if probabilities do not sum to 1
    rho_sum = sum(rho.values())
    if abs(rho_sum - 1.0) > 1e-6:
        print(f"  WARNING: scenario probabilities sum to {rho_sum:.4f}, not 1.0.")

    params["d"] = d
    params["rho"] = rho

    # ── Sheet 3: Vehicles ────────────────────────────────────────────────────
    ws = wb["Vehicles"]
    q = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            k = int(row[0])
            q[k] = float(row[1])

    if len(q) != params["nK"]:
        raise ValueError(
            f"Vehicles sheet has {len(q)} rows but nK={params['nK']}. "
            f"Please add or remove rows to match."
        )
    params["q"] = q

    # ── Sheet 4: Farmers ─────────────────────────────────────────────────────
    ws = wb["Farmers"]
    r_price = {}
    alpha = {}
    a0 = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            i = int(row[0])
            r_price[i] = float(row[1])
            alpha[i] = float(row[2])
            a0[i] = float(row[3])

    if len(r_price) != params["nF"]:
        raise ValueError(
            f"Farmers sheet has {len(r_price)} rows but nF={params['nF']}. "
            f"Please add or remove rows to match."
        )
    params["r_price"] = r_price
    params["alpha"] = alpha
    params["a0"] = a0

    # ── Sheet 5: Harvest ─────────────────────────────────────────────────────
    ws = wb["Harvest"]
    g = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            i = int(row[0])
            t = int(row[1])
            g[i, t] = float(row[2])

    expected = params["nF"] * params["nT"]
    if len(g) != expected:
        raise ValueError(
            f"Harvest sheet has {len(g)} rows but nF × nT = {expected}. "
            f"Every farmer-period combination must have exactly one row."
        )
    params["g"] = g

    # ── Sheet 6: Coordinates (optional) ─────────────────────────────────────
    coords = {}
    if "Coordinates" in wb.sheetnames:
        ws = wb["Coordinates"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and row[1] is not None:
                node = int(row[0])
                coords[node] = (float(row[1]), float(row[2]))

    params["coords_from_excel"] = coords  # may be empty → auto-generate below

    return params, N, K, T, S


# =============================================================================
# 2.  BUILD INSTANCE
# =============================================================================

print(f"\nReading parameters from: {PARAMS_FILE}")
params, N, K, T, S = load_parameters(PARAMS_FILE)

nF = params["nF"]
nK = params["nK"]
nT = params["nT"]
nS = params["nS"]
nO = nK

T0 = list(range(0, nT + 1))
O = list(range(nF + 1, nF + nO + 1))
V = list(range(0, nF + nO + 1))
HUB = 0

rho = params["rho"]
d = params["d"]
q = params["q"]
r_price = params["r_price"]
alpha = params["alpha"]
a0 = params["a0"]
g = params["g"]

originOf = {k: nF + k for k in K}

# ── Coordinates ──────────────────────────────────────────────────────────────
# If the Coordinates sheet was populated, use those values.
# Otherwise, auto-generate from seed=42 (reproducible fallback).
coords_excel = params["coords_from_excel"]

if len(coords_excel) == len(V):
    coords = coords_excel
    print("  Coordinates loaded from Excel.")
else:
    import random

    random.seed(42)
    # Consume the same random calls the original code used before coords
    # (r_price_crate, alpha, a0_crate, g_crate generations) so that
    # coordinates remain identical to the original random instance.
    _ = {i: random.uniform(2.50, 3.50) for i in N}
    _ = {i: random.uniform(4.00, 7.00) for i in N}
    _ = {i: random.randint(0, 49) for i in N}
    _ = {(i, t): random.randint(50, 149) * (1 + 0.1 * random.gauss(0, 1))
         for i in N for t in T}
    coords = {0: (38.40, 27.14)}
    for i in N:
        coords[i] = (37.50 + random.random(), 27.80 + random.random() * 0.6)
    for o in O:
        b = random.choice(N)
        coords[o] = (coords[b][0] + random.gauss(0, 0.05),
                     coords[b][1] + random.gauss(0, 0.05))
    print("  Coordinates auto-generated from seed=42.")

def dist_km(u, v):
    if u == v: return 0.0
    dlat = (coords[v][0]-coords[u][0]) * 111.0
    dlon = (coords[v][1]-coords[u][1]) * 111.0 * math.cos(math.radians(38.0))
    return math.sqrt(dlat**2 + dlon**2)

cost_per_km = 12/100 * 1.29
c = {(u, v, k): round(dist_km(u,v)*cost_per_km, 4)
     for u in V for v in V for k in K}

def m_val(i, tau, t, s):
    return round(r_price[i] * (1 - d[s])**(t - tau), 6) #C29
def ub_cohort_val(i, tau, t, s):
    if tau > t: return 0.0
    if tau == 0: return round(a0[i] * (1 - d[s])**(t - 1), 6)
    return round(g[i, tau] * (1 - d[s])**(t - tau), 6) #C30
def ub_total_val(i, t, s):
    return round(sum(ub_cohort_val(i, tau, t, s) for tau in T0 if tau <= t), 6) #C31

m         = {(i,tau,t,s): m_val(i,tau,t,s)
             for i in N for tau in T0 for t in T for s in S if tau <= t}
ub_cohort = {(i,tau,t,s): ub_cohort_val(i,tau,t,s)
             for i in N for tau in T0 for t in T for s in S if tau <= t}
ub_total  = {(i,t,s): ub_total_val(i,t,s)
             for i in N for t in T for s in S}

def solve_with_cplex():
    from docplex.mp.model import Model
    mdl = Model(name="Perishable_VRP_kg")

    X = {(u,v,k,t): mdl.binary_var(name=f"X_{u}_{v}_{k}_{t}")
         for u in V for v in V for k in K for t in T if u!=v}
    Z     = {(i,k,t): mdl.binary_var(name=f"Z_{i}_{k}_{t}") for i in N for k in K for t in T}
    delta = {(i,t): mdl.binary_var(name=f"delta_{i}_{t}") for i in N for t in T}
    W     = {(k,t): mdl.binary_var(name=f"W_{k}_{t}") for k in K for t in T}
    B   = {(i,tau,t,s): mdl.continuous_var(lb=0) for i in N for tau in T0 for t in T for s in S if tau<=t}
    P   = {(i,tau,t,s): mdl.continuous_var(lb=0) for i in N for tau in T0 for t in T for s in S if tau<=t}
    Inv = {(i,tau,t,s): mdl.continuous_var(lb=0) for i in N for tau in T0 for t in T for s in S if tau<=t}
    Q   = {(i,t,s): mdl.continuous_var(lb=0) for i in N for t in T for s in S}
    L   = {(i,k,t,s): mdl.continuous_var(lb=0) for i in N for k in K for t in T for s in S}
    F   = {(u,v,k,t,s): mdl.continuous_var(lb=0)
           for u in V for v in V for k in K for t in T for s in S if u!=v}
    M_aux = {(i,k,t): mdl.continuous_var(lb=0) for i in N for k in K for t in T}

    revenue = mdl.sum(rho[s]*m[i,tau,t,s]*P[i,tau,t,s] for i in N for t in T for s in S for tau in T0 if tau<=t)
    travel  = mdl.sum(c[u,v,k]*X[u,v,k,t] for u in V for v in V for k in K for t in T if u!=v)
    service = mdl.sum(alpha[i]*Z[i,k,t] for i in N for k in K for t in T)
    mdl.maximize(revenue - travel - service)

    for i in N:
        for s in S:
            mdl.add_constraint(B[i,0,1,s] == a0[i]) #C2

    for i in N:
        for t in T:
            for s in S:
                if t > 1:
                    mdl.add_constraint(B[i,0,t,s] == (1-d[s])*Inv[i,0,t-1,s]) #C3

    for i in N:
        for t in T:
            for s in S:
                mdl.add_constraint(B[i,t,t,s] == g[i,t]) #C4

    for i in N:
        for t in T:
            for s in S:
                for tau in T:
                    if 1 <= tau < t:
                        mdl.add_constraint(B[i,tau,t,s]==(1-d[s])*Inv[i,tau,t-1,s]) #C5
    for i in N:
        for t in T:
            mdl.add_constraint(delta[i,t] <= mdl.sum(Z[i,k,t] for k in K)) #C6

    for i in N:
        for k in K:
            for t in T:
                mdl.add_constraint(delta[i,t] >= Z[i,k,t]) #C7
    for i in N:
        for t in T:
            for s in S:
                for tau in T0:
                    if tau <= t:
                        ub = ub_cohort[i,tau,t,s]
                        mdl.add_constraint(P[i,tau,t,s] <= B[i,tau,t,s]) #C8
                        mdl.add_constraint(P[i,tau,t,s] <= ub*delta[i,t]) #C9
                        mdl.add_constraint(P[i,tau,t,s] >= B[i,tau,t,s]-ub*(1-delta[i,t])) #C10
                        mdl.add_constraint(Inv[i,tau,t,s]==B[i,tau,t,s]-P[i,tau,t,s]) #C11

    for i in N:
        for t in T:
            for s in S:
                mdl.add_constraint(Q[i, t, s] == mdl.sum(P[i, tau, t, s] for tau in T0 if tau <= t)) #C12
                mdl.add_constraint(mdl.sum(L[i, k, t, s] for k in K) == Q[i, t, s]) #C13

    for i in N:
        for k in K:
            for t in T:
                for s in S:
                    mdl.add_constraint(L[i, k, t, s] <= q[k] * Z[i, k, t]) #C14
                    mdl.add_constraint(L[i, k, t, s] <= ub_total[i, t, s] * Z[i, k, t]) #C15
    for u in V:
        for v in V:
            if u == v: continue
            for k in K:
                for t in T:
                    for s in S:
                        mdl.add_constraint(F[u, v, k, t, s] <= q[k] * X[u, v, k, t]) #C16


    for k in K:
        ok = originOf[k]
        for t in T:
            for s in S:
                mdl.add_constraint(mdl.sum(F[ok,v,k,t,s] for v in V if v!=ok)==0) #C17
                mdl.add_constraint(mdl.sum(F[u,ok,k,t,s] for u in V if u!=ok)==0)  #C18
                mdl.add_constraint(mdl.sum(F[u,HUB,k,t,s] for u in V if u!=HUB)==mdl.sum(L[i,k,t,s] for i in N)) #C20
                mdl.add_constraint(mdl.sum(F[HUB,v,k,t,s] for v in V if v!=HUB)==0) #C21
    for i in N:
        for k in K:
            for t in T:
                for s in S:
                    mdl.add_constraint(mdl.sum(F[i,v,k,t,s] for v in V if v!=i)-mdl.sum(F[u,i,k,t,s] for u in V if u!=i)==L[i,k,t,s]) #C19
    for k in K:
        ok = originOf[k]
        for t in T:
            mdl.add_constraint(mdl.sum(X[ok,v,k,t] for v in V if v!=ok)==W[k,t]) #C22
            mdl.add_constraint(mdl.sum(X[u,HUB,k,t] for u in V if u!=HUB)==W[k,t]) #C23
            mdl.add_constraint(mdl.sum(X[u,ok,k,t] for u in V if u!=ok)==0) #C24
            mdl.add_constraint(mdl.sum(X[HUB,v,k,t] for v in V if v!=HUB)==0) #C25
    for i in N:
        for k in K:
            for t in T:
                mdl.add_constraint(mdl.sum(X[u,i,k,t] for u in V if u!=i)==Z[i,k,t])  #26
                mdl.add_constraint(mdl.sum(X[i,v,k,t] for v in V if v!=i)==Z[i,k,t])  #27
    for i in N:
        for j in N:
            if i==j: continue
            for k in K:
                for t in T:
                    mdl.add_constraint(M_aux[i,k,t]-M_aux[j,k,t]+nF*X[i,j,k,t]<=nF-1+nF*(2-Z[i,k,t]-Z[j,k,t]))  #28

    for i in N:
        for t in T:
            for s in S:
                for tau in T0:
                    if tau <= t:
                        ub = ub_cohort[i, tau, t, s]
                        mdl.add_constraint(B[i, tau, t, s] <= ub)
                        mdl.add_constraint(P[i, tau, t, s] <= ub)
                        mdl.add_constraint(Inv[i, tau, t, s] <= ub)
                        mdl.add_constraint(Q[i, t, s] <= ub_total[i, t, s])

    t0 = time.time()
    sol = mdl.solve(log_output=True)
    elapsed = time.time() - t0
    status = str(mdl.solve_details.status) if mdl.solve_details else "Unknown"
    obj = mdl.objective_value if sol else None

    # ── Decompose objective into its three components ─────────────────────────
    if sol:
        rev_val = sum(rho[s] * m[i, tau, t, s] * P[i, tau, t, s].solution_value
                      for i in N for t in T for s in S
                      for tau in T0 if tau <= t)
        trav_val = sum(c[u, v, k] * X[u, v, k, t].solution_value
                       for u in V for v in V for k in K for t in T if u != v)
        serv_val = sum(alpha[i] * Z[i, k, t].solution_value
                       for i in N for k in K for t in T)
        rev_val = round(rev_val, 4)
        trav_val = round(trav_val, 4)
        serv_val = round(serv_val, 4)
    else:
        rev_val = trav_val = serv_val = None

    return mdl, status, obj, elapsed, X, Z, delta, W, B, P, Inv, Q, L, F, M_aux, \
        rev_val, trav_val, serv_val


# =============================================================================
# 2.  EXCEL EXPORT
# =============================================================================

def export_to_excel(status, obj, elapsed, X, Z, delta, W, B, P, Inv, Q, L, F, M_aux,
                    rev_val, trav_val, serv_val, filename="VRP_results_kg2.xlsx"):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    def val(v):
        if hasattr(v, "varValue"):        return v.varValue
        if hasattr(v, "solution_value"):  return v.solution_value
        if hasattr(v, "X"):               return v.X
        return float(v)

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    center      = Alignment(horizontal="center")

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center

    # ── Summary ───────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    style_header(ws, ["Item", "Value"])
    for r, (item, value) in enumerate([
        ("Solver",              SOLVER.upper()),
        ("Status",              status),
        ("Objective (EUR)", round(obj, 4) if obj is not None else "N/A"),
        ("  Revenue (EUR)", rev_val if rev_val is not None else "N/A"),  # new
        ("  Travel cost (EUR)", trav_val if trav_val is not None else "N/A"),  # new
        ("  Service cost (EUR)", serv_val if serv_val is not None else "N/A"),  # new
        ("Wall time (s)", round(elapsed, 2)),
        ("Farmers (nF)",        nF),
        ("Vehicles (nK)",       nK),
        ("Periods (nT)",        nT),
        ("Scenarios (nS)",      nS),
        ("Vehicle capacity",    f"{q[1]} kg"),
    ], start=2):
        ws.cell(row=r, column=1, value=item)
        ws.cell(row=r, column=2, value=value)

    # ── Parameters ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("Parameters")
    style_header(ws, ["Parameter", "Index", "Value", "Description"])
    r = 2
    for i in N:
        ws.cell(r, 1, "r_price"); ws.cell(r, 2, f"i={i}")
        ws.cell(r, 3, r_price[i]); ws.cell(r, 4, "EUR/kg, fresh unit price"); r += 1
    for i in N:
        ws.cell(r, 1, "alpha"); ws.cell(r, 2, f"i={i}")
        ws.cell(r, 3, alpha[i]); ws.cell(r, 4, "EUR/visit service cost"); r += 1
    for i in N:
        ws.cell(r, 1, "a0"); ws.cell(r, 2, f"i={i}")
        ws.cell(r, 3, a0[i]); ws.cell(r, 4, "initial inventory (kg)"); r += 1
    for (i, t) in sorted(g.keys()):
        ws.cell(r, 1, "g"); ws.cell(r, 2, f"i={i}, t={t}")
        ws.cell(r, 3, g[i, t]); ws.cell(r, 4, "harvest quantity (kg)"); r += 1
    for k in K:
        ws.cell(r, 1, "q"); ws.cell(r, 2, f"k={k}")
        ws.cell(r, 3, q[k]); ws.cell(r, 4, "vehicle capacity (kg)"); r += 1
    for s in S:
        ws.cell(r, 1, "rho"); ws.cell(r, 2, f"s={s}")
        ws.cell(r, 3, rho[s]); ws.cell(r, 4, "scenario probability"); r += 1
    for s in S:
        ws.cell(r, 1, "d"); ws.cell(r, 2, f"s={s}")
        ws.cell(r, 3, d[s]); ws.cell(r, 4, "deterioration rate"); r += 1

    if obj is None:
        wb.save(filename)
        print(f"  Exported (no solution) -> {filename}")
        return

    # ── X: route arcs ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("X_route_arcs")
    style_header(ws, ["u", "v", "k", "t", "X"])
    r = 2
    for (u, v, k, t), var in X.items():
        v_val = round(val(var), 6)
        if v_val > 1e-6:
            ws.cell(r, 1, u); ws.cell(r, 2, v); ws.cell(r, 3, k)
            ws.cell(r, 4, t); ws.cell(r, 5, v_val)
            r += 1

    # ── Z: vehicle-farmer assignment ──────────────────────────────────────────
    ws = wb.create_sheet("Z_vehicle_farmer")
    style_header(ws, ["i (farmer)", "k (vehicle)", "t (period)", "Z"])
    r = 2
    for (i, k, t), var in Z.items():
        ws.cell(r, 1, i); ws.cell(r, 2, k); ws.cell(r, 3, t)
        ws.cell(r, 4, round(val(var), 6)); r += 1

    # ── delta: farmer visit ────────────────────────────────────────────────────
    ws = wb.create_sheet("delta_farmer_visit")
    style_header(ws, ["i (farmer)", "t (period)", "delta"])
    r = 2
    for (i, t), var in delta.items():
        ws.cell(r, 1, i); ws.cell(r, 2, t)
        ws.cell(r, 3, round(val(var), 6)); r += 1

    # ── W: vehicle active ─────────────────────────────────────────────────────
    ws = wb.create_sheet("W_vehicle_active")
    style_header(ws, ["k (vehicle)", "t (period)", "W"])
    r = 2
    for (k, t), var in W.items():
        ws.cell(r, 1, k); ws.cell(r, 2, t)
        ws.cell(r, 3, round(val(var), 6)); r += 1

    # ── B: available inventory ────────────────────────────────────────────────
    ws = wb.create_sheet("B_available_inv")
    style_header(ws, ["i (farmer)", "tau (cohort)", "t (period)", "s (scenario)", "B (kg)"])
    r = 2
    for (i, tau, t, s), var in B.items():
        ws.cell(r, 1, i); ws.cell(r, 2, tau); ws.cell(r, 3, t); ws.cell(r, 4, s)
        ws.cell(r, 5, round(val(var), 4)); r += 1

    # ── P: picked up ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("P_pickup")
    style_header(ws, ["i (farmer)", "tau (cohort)", "t (period)", "s (scenario)", "P (kg)"])
    r = 2
    for (i, tau, t, s), var in P.items():
        ws.cell(r, 1, i); ws.cell(r, 2, tau); ws.cell(r, 3, t); ws.cell(r, 4, s)
        ws.cell(r, 5, round(val(var), 4)); r += 1

    # ── Inv: leftover inventory ───────────────────────────────────────────────
    ws = wb.create_sheet("Inv_leftover")
    style_header(ws, ["i (farmer)", "tau (cohort)", "t (period)", "s (scenario)", "Inv (kg)"])
    r = 2
    for (i, tau, t, s), var in Inv.items():
        ws.cell(r, 1, i); ws.cell(r, 2, tau); ws.cell(r, 3, t); ws.cell(r, 4, s)
        ws.cell(r, 5, round(val(var), 4)); r += 1

    # ── Q: total collected ────────────────────────────────────────────────────
    ws = wb.create_sheet("Q_total_collected")
    style_header(ws, ["i (farmer)", "t (period)", "s (scenario)", "Q (kg)"])
    r = 2
    for (i, t, s), var in Q.items():
        ws.cell(r, 1, i); ws.cell(r, 2, t); ws.cell(r, 3, s)
        ws.cell(r, 4, round(val(var), 4)); r += 1

    # ── L: load per vehicle ───────────────────────────────────────────────────
    ws = wb.create_sheet("L_load")
    style_header(ws, ["i (farmer)", "k (vehicle)", "t (period)", "s (scenario)", "L (kg)"])
    r = 2
    for (i, k, t, s), var in L.items():
        ws.cell(r, 1, i); ws.cell(r, 2, k); ws.cell(r, 3, t); ws.cell(r, 4, s)
        ws.cell(r, 5, round(val(var), 4)); r += 1

    # ── F: flow (nonzero only) ────────────────────────────────────────────────
    ws = wb.create_sheet("F_flow")
    style_header(ws, ["u", "v", "k", "t", "s", "F (kg)"])
    r = 2
    for (u, v, k, t, s), var in F.items():
        f_val = round(val(var), 4)
        if f_val > 1e-6:
            ws.cell(r, 1, u); ws.cell(r, 2, v); ws.cell(r, 3, k)
            ws.cell(r, 4, t); ws.cell(r, 5, s); ws.cell(r, 6, f_val)
            r += 1

    # ── M_aux: MTZ subtour ────────────────────────────────────────────────────
    ws = wb.create_sheet("Maux_MTZ")
    style_header(ws, ["i (farmer)", "k (vehicle)", "t (period)", "Maux"])
    r = 2
    for (i, k, t), var in M_aux.items():
        ws.cell(r, 1, i); ws.cell(r, 2, k); ws.cell(r, 3, t)
        ws.cell(r, 4, round(val(var), 4)); r += 1

    wb.save(filename)
    print(f"\n  Results exported -> {filename}")


# =============================================================================
# 3.  RESULT PRINTER
# =============================================================================

def print_results(status, obj, elapsed, X, Z, delta, W, B, P, Q, L,
                  rev_val, trav_val, serv_val):
    print("\n" + "="*60)
    print(f"  SOLVER   : {SOLVER.upper()}")
    print(f"  STATUS   : {status}")
    print(f"  OBJECTIVE: {round(obj, 4) if obj is not None else 'N/A'} EUR")
    if obj is not None:  # new
        print(f"    Revenue      : {rev_val} EUR")  # new
        print(f"    Travel cost  : {trav_val} EUR")  # new
        print(f"    Service cost : {serv_val} EUR")  # new
    print(f"  WALL TIME: {round(elapsed,2)} s")
    print("="*60)

    if obj is None:
        print("  No feasible solution found.")
        return

    def val(v):
        """Extract numeric value regardless of solver variable type."""
        if hasattr(v, "varValue"):        return v.varValue   # PuLP
        if hasattr(v, "solution_value"):  return v.solution_value  # CPLEX
        if hasattr(v, "X"):               return v.X           # Gurobi
        return float(v)

    print("\n--- Active routes (X=1) ---")
    for t in T:
        print(f"  Period {t}:")
        for k in K:
            route_arcs = [(u,v) for (u,v,kk,tt) in X
                          if kk==k and tt==t and val(X[u,v,k,t]) > 0.5]
            if route_arcs:
                print(f"    Vehicle {k}: {route_arcs}")

    print("\n--- Farmer visits (delta=1) ---")
    for t in T:
        visited = [i for i in N if val(delta[i,t]) > 0.5]
        print(f"  Period {t}: farmers {visited}")

    print("\n--- Collected quantities Q[i,t,s] (kg) ---")
    for t in T:
        for s in S:
            row_q = {i: round(val(Q[i,t,s]),2) for i in N}
            print(f"  Period {t}, Scenario {s}: {row_q}")


# =============================================================================
# 4.  MAIN ENTRY POINT
# =============================================================================

if __name__ == "__main__":
    print(f"\nBuilding model (kg) | Solver: {SOLVER.upper()}")
    print(f"Instance: {nF} farmers | {nK} vehicles | {nT} periods | {nS} scenarios")
    print(f"Nodes: {len(V)}  (hub=0, farmers=1..{nF}, origins={nF+1}..{nF+nO})")
    for i in N:
        print(f"  Farmer {i}: r={r_price[i]} EUR/kg, alpha={alpha[i]} EUR/visit, "
              f"a0={a0[i]} kg, harvests={[g[i,t] for t in T]}")
    print()

    if SOLVER == "cplex":
        mdl, status, obj, elapsed, X, Z, delta, W, B, P, Inv, Q, L, F, M_aux, \
        rev_val, trav_val, serv_val = solve_with_cplex()
    else:
        raise ValueError(f"Unknown solver '{SOLVER}'. Only 'cplex' is implemented in this file.")

    print_results(status, obj, elapsed, X, Z, delta, W, B, P, Q, L,
                  rev_val, trav_val, serv_val)
    export_to_excel(status, obj, elapsed, X, Z, delta, W, B, P, Inv, Q, L, F, M_aux,
                    rev_val, trav_val, serv_val)