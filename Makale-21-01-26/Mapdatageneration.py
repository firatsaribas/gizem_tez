"""
OPL Excel → MapData CSV Converter
===================================
Reads OPL_Turkey_Datasets.xlsx and exports one CSV per dataset sheet
in the format: Ad, Enlem, Boylam, Type

Usage:
  python excel_to_mapcsv.py                              # converts all sheets
  python excel_to_mapcsv.py --sheet Örnek_1              # single sheet
  python excel_to_mapcsv.py --sheet Temel_Durum --out MyMap.csv

Output files (one per sheet):
  MapData_Temel_Durum.csv
  MapData_Örnek_1.csv
  MapData_Örnek_2.csv
  MapData_Örnek_3.csv
  MapData_Örnek_4.csv
"""
"""
import pandas as pd
import openpyxl
import argparse
import os

# Sheet names to process (skip the summary sheet)
DATASET_SHEETS = ["Temel_Durum", "Örnek_1", "Örnek_2", "Örnek_3", "Örnek_4"]

# Column positions in the Excel sheet (1-indexed)
# Suppliers: cols A=1 (No), B=2 (Lat), C=3 (Lon)
# Hubs:      cols E=5 (No), F=6 (Lat), G=7 (Lon)
# Depots:    cols I=9 (No), J=10 (Lat), K=11 (Lon)
COL_SUPPLIER = (1, 2, 3)   # (No, Lat, Lon)
COL_HUB      = (5, 6, 7)
COL_DEPOT    = (9, 10, 11)

# Type labels (must match your MapData CSV)
TYPE_SUPPLIER = "Tedarikçi"
TYPE_HUB      = "Hub"
TYPE_DEPOT    = "Depo"

# Data starts at row 4 (rows 1-3 are title/headers)
DATA_START_ROW = 4


def extract_sheet(ws):
    #Extract all nodes from one Excel sheet, grouped by type (Suppliers → Hubs → Depots).
    suppliers, hubs, depots = [], [], []

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        # Suppliers
        no, lat, lon = row[COL_SUPPLIER[0]-1], row[COL_SUPPLIER[1]-1], row[COL_SUPPLIER[2]-1]
        if no and lat and lon and isinstance(lat, float):
            suppliers.append({"Ad": no, "Enlem": lat, "Boylam": lon, "Type": TYPE_SUPPLIER})

        # Hubs
        no, lat, lon = row[COL_HUB[0]-1], row[COL_HUB[1]-1], row[COL_HUB[2]-1]
        if no and lat and lon and isinstance(lat, float):
            hubs.append({"Ad": no, "Enlem": lat, "Boylam": lon, "Type": TYPE_HUB})

        # Depots
        no, lat, lon = row[COL_DEPOT[0]-1], row[COL_DEPOT[1]-1], row[COL_DEPOT[2]-1]
        if no and lat and lon and isinstance(lat, float):
            depots.append({"Ad": no, "Enlem": lat, "Boylam": lon, "Type": TYPE_DEPOT})

    # Group: all suppliers first, then hubs, then depots — matching MapData CSV format
    return suppliers + hubs + depots


def convert(excel_path, sheet_name=None, output_path=None):
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    sheets = [sheet_name] if sheet_name else DATASET_SHEETS

    for sheet in sheets:
        if sheet not in wb.sheetnames:
            print(f"  WARNING: Sheet '{sheet}' not found, skipping.")
            continue

        ws = wb[sheet]
        rows = extract_sheet(ws)

        if not rows:
            print(f"  WARNING: No data found in sheet '{sheet}', skipping.")
            continue

        df = pd.DataFrame(rows, columns=["Ad", "Enlem", "Boylam", "Type"])

        # Output path
        if output_path and sheet_name:
            out = output_path
        else:
            base = os.path.splitext(excel_path)[0]
            out = f"MapData_{sheet}.csv"

        df.to_csv(out, index=False, encoding="utf-8-sig")  # utf-8-sig for Excel compatibility

        counts = df["Type"].value_counts()
        print(f"  Saved: {out}")
        print(f"    Tedarikçi: {counts.get(TYPE_SUPPLIER, 0)}  "
              f"Hub: {counts.get(TYPE_HUB, 0)}  "
              f"Depo: {counts.get(TYPE_DEPOT, 0)}  "
              f"Toplam: {len(df)}")


def main():
    parser = argparse.ArgumentParser(description="OPL Excel to MapData CSV converter")
    parser.add_argument("excel", nargs="?",
                        default="OPL_Turkey_Datasets_v5.xlsx",
                        help="Input Excel file (default: OPL_Turkey_Datasets_v5.xlsx)")
    parser.add_argument("--sheet", help="Single sheet name to convert (default: all)")
    parser.add_argument("--out",   help="Output CSV path (only used with --sheet)")
    args = parser.parse_args()

    if not os.path.exists(args.excel):
        print(f"ERROR: File not found: {args.excel}")
        return

    print(f"Reading: {args.excel}")
    convert(args.excel, sheet_name=args.sheet, output_path=args.out)
    print("Done.")


if __name__ == "__main__":
    main()
"""
"""
MapData CSV Merger
===================
Merges multiple MapData CSV files into a single CSV file.

Usage:
  python merge_mapcsv.py                        # merges all MapData_*.csv in current folder
  python merge_mapcsv.py --out AllSamples.csv   # custom output name
"""

import pandas as pd
import glob
import os
import argparse

def merge_csv(output_path):
    # Find all MapData CSV files in the current folder
    csv_files = sorted(glob.glob("MapData_*.csv"))

    if not csv_files:
        print("No MapData_*.csv files found in the current folder.")
        return

    print(f"Found {len(csv_files)} files to merge:")

    all_dfs = []
    for f in csv_files:
        df = pd.read_csv(f, encoding="utf-8-sig")
        df["Dataset"] = os.path.splitext(f)[0].replace("MapData_", "")  # adds source column
        print(f"  {f}  →  {len(df)} rows")
        all_dfs.append(df)

    merged = pd.concat(all_dfs, ignore_index=True)
    merged.to_csv(output_path, index=False, encoding="utf-8-sig")

    print(f"\nSaved: {output_path}")
    print(f"  Total rows: {len(merged)}")
    print(f"  Tedarikçi: {len(merged[merged['Type']=='Tedarikçi'])}")
    print(f"  Hub:        {len(merged[merged['Type']=='Hub'])}")
    print(f"  Depo:       {len(merged[merged['Type']=='Depo'])}")

def main():
    parser = argparse.ArgumentParser(description="Merge MapData CSV files into one")
    parser.add_argument("--out", default="MapData_All.csv", help="Output file name")
    args = parser.parse_args()
    merge_csv(args.out)

if __name__ == "__main__":
    main()