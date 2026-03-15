import random
import pandas as pd
from geopy.distance import geodesic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG — change only here
# ============================================================
NUM_DEPOTS = 40    # total number of depots
NUM_HUBS   = 4     # total number of hubs
TARGET     = 500  # total number of routes to generate

NOKTALAR_FILE = r"C:\Users\gizem\OneDrive\Belgeler\GitHub\gizem_tez\Makale-21-01-26\Noktalar150suppliers.csv"

FUEL_PRICE_PER_LITRE = 1.29

# Capacity and fuel consumption by number of depots visited
CAPACITY_BY_NUM_DEPOTS  = {3: 750,  4: 900,  5: 1200, 6: 1400}
FUEL_CONSUMPTION_BY_NUM = {3: 25,   4: 30,   5: 38,   6: 38}   # litres per 100 km

RETURN_TO_HUB = False
# ============================================================

# Derived — do not change manually
ALL_DEPOTS = list(range(1, NUM_DEPOTS + 1))
ALL_HUBS   = list(range(1, NUM_HUBS   + 1))

# =========================
# 0) INPUT: base 30 routes
# =========================
route_to_am_30 = {
    1:2, 2:1, 3:1, 4:2, 5:3, 6:3, 7:2, 8:3, 9:2, 10:1,
    11:2, 12:3, 13:2, 14:1, 15:3, 16:2, 17:2, 18:2, 19:1, 20:3,
    21:2, 22:3, 23:3, 24:2, 25:1, 26:1, 27:3, 28:1, 29:1, 30:1
}

route_to_depots_30 = {
    1:[25,24,14,13], 2:[16,17,23,20], 3:[10,22,27,30], 4:[28,26,19],
    5:[11,12,29],    6:[8,18,21,15,9],7:[19,20,14,7,2], 8:[17,15,21,25],
    9:[22,8,6,5,1],  10:[23,13,10,11],11:[29,26,24,16,4,3],12:[9,18,30,28,27],
    13:[30,23,5,4,2],14:[12,15,21,25,28],15:[7,9,10,16,22],16:[18,19,17,14,12],
    17:[29,24,8,1,3],18:[27,20,7,6,4,2],19:[7,6,5,3,1],  20:[20,21,22,24,26,28],
    21:[23,8,16,13,11],22:[6,5,4,3,2,1],23:[17,30,14,9,10,12],24:[18,19,25,26,27,29],
    25:[11,13,15,20,22,26],26:[10,16,15,27],27:[7,13,14,17,28],28:[9,12,11],
    29:[8,19,23,25,24,21],30:[18,30,29]
}

# =========================
# 1) Route generation helpers
# =========================
def normalize_route(depots):
    seen = set(); out = []
    for d in depots:
        if d not in seen:
            out.append(d); seen.add(d)
    return out

def mutate_route(depots, hub, min_len=3, max_len=6):
    depots = depots[:]
    op = random.choice(["swap", "add", "remove", "replace"])
    if op == "swap" and len(depots) >= 2:
        i, j = random.sample(range(len(depots)), 2)
        depots[i], depots[j] = depots[j], depots[i]
    elif op == "add" and len(depots) < max_len:
        cand = [d for d in ALL_DEPOTS if d not in depots]
        if cand:
            depots.insert(random.randrange(len(depots) + 1), random.choice(cand))
    elif op == "remove" and len(depots) > min_len:
        depots.pop(random.randrange(len(depots)))
    elif op == "replace":
        cand = [d for d in ALL_DEPOTS if d not in depots]
        if cand and depots:
            depots[random.randrange(len(depots))] = random.choice(cand)
    depots = normalize_route(depots)
    if len(depots) < min_len:
        cand = [d for d in ALL_DEPOTS if d not in depots]
        random.shuffle(cand)
        depots += cand[:(min_len - len(depots))]
    if len(depots) > max_len:
        depots = depots[:max_len]
    return depots

def random_route(min_len=3, max_len=6):
    return random.sample(ALL_DEPOTS, random.randint(min_len, max_len))

def merge_routes(a, b, min_len=3, max_len=6):
    merged = normalize_route(a + b)
    if len(merged) < min_len:
        cand = [d for d in ALL_DEPOTS if d not in merged]
        random.shuffle(cand)
        merged += cand[:(min_len - len(merged))]
    if len(merged) > max_len:
        merged = merged[:max_len]
    return merged

def signature(hub, depots):
    return (hub, tuple(sorted(depots)))

def expand_routes_to_target(route_to_am, route_to_depots,
                             target=1000, seed=42, min_len=3, max_len=6,
                             n_mutate=None, n_random=None, n_merge=None):
    random.seed(seed)
    if n_mutate is None: n_mutate = round(target * 0.370)
    if n_random is None: n_random = round(target * 0.405)
    if n_merge  is None: n_merge  = round(target * 0.195)
    new_am  = dict(route_to_am)
    new_dep = {r: normalize_route(ds) for r, ds in route_to_depots.items()}
    used    = set(signature(new_am[r], new_dep[r]) for r in new_dep)
    next_id = max(new_dep.keys()) + 1

    base_routes = list(new_dep.keys())
    for _ in range(n_mutate):
        if next_id > target: break
        r0 = random.choice(base_routes)
        hub = new_am[r0]
        depots = mutate_route(new_dep[r0], hub, min_len, max_len)
        sig = signature(hub, depots)
        if sig in used: continue
        new_am[next_id] = hub; new_dep[next_id] = depots; used.add(sig); next_id += 1

    for _ in range(n_random):
        if next_id > target: break
        hub = random.choice(ALL_HUBS)
        depots = random_route(min_len, max_len)
        sig = signature(hub, depots)
        if sig in used: continue
        new_am[next_id] = hub; new_dep[next_id] = depots; used.add(sig); next_id += 1

    by_hub = {h: [] for h in ALL_HUBS}
    for r in new_dep: by_hub[new_am[r]].append(r)
    for _ in range(n_merge):
        if next_id > target: break
        hub = random.choice(ALL_HUBS)
        if len(by_hub[hub]) < 2: continue
        r1, r2 = random.sample(by_hub[hub], 2)
        depots = merge_routes(new_dep[r1], new_dep[r2], min_len, max_len)
        sig = signature(hub, depots)
        if sig in used: continue
        new_am[next_id] = hub; new_dep[next_id] = depots; used.add(sig)
        by_hub[hub].append(next_id); next_id += 1

    attempts = 0
    while next_id <= target:
        attempts += 1
        if attempts > 100000:
            print(f"WARNING: Only generated {next_id - 1} unique routes."); break
        if random.random() < 0.5:
            r0 = random.choice(list(new_dep.keys()))
            hub = new_am[r0]
            depots = mutate_route(new_dep[r0], hub, min_len, max_len)
        else:
            hub = random.choice(ALL_HUBS)
            depots = random_route(min_len, max_len)
        sig = signature(hub, depots)
        if sig in used: continue
        new_am[next_id] = hub; new_dep[next_id] = depots; used.add(sig); next_id += 1

    return new_am, new_dep

# =========================
# 2) Generate routes
# =========================
print("Rotalar üretiliyor...")
# n_mutate / n_random / n_merge scale automatically with TARGET
# ratios: 37% / 40.5% / 19.5%  (preserving original proportions)
n_mutate = round(TARGET * 0.370)
n_random = round(TARGET * 0.405)
n_merge  = round(TARGET * 0.195)

route_to_am_out, route_to_depots_out = expand_routes_to_target(
    route_to_am_30, route_to_depots_30,
    target=TARGET, seed=42, min_len=3, max_len=6,
    n_mutate=n_mutate, n_random=n_random, n_merge=n_merge
)
print(f"Toplam üretilen rota sayısı: {len(route_to_depots_out)}")

# =========================
# 3) Build route points list (shared by CSV + analysis)
# =========================
def build_route_points(hub_id, depots):
    pts = [f"Bölge birliği {hub_id}"] + [f"Depo {d}" for d in depots]
    if RETURN_TO_HUB:
        pts.append(f"Bölge birliği {hub_id}")
    return pts

all_route_points = {}   # rid -> list of point name strings
for rid in range(1, TARGET + 1):
    all_route_points[rid] = build_route_points(
        route_to_am_out[rid], route_to_depots_out[rid]
    )

# =========================
# 4) Save CSV
# =========================
rotalarr_lines = [",".join([f"Rota {rid}"] + all_route_points[rid])
                  for rid in range(1, TARGET + 1)]
csv_out = f"rotalarr_{TARGET}_generated.csv"
with open(csv_out, "w", encoding="utf-8-sig") as f:
    f.write("\n".join(rotalarr_lines))
print(f"CSV yazıldı: {csv_out}")

# =========================
# 5) Load Noktalar & compute cost/capacity
# =========================
print("Noktalar yükleniyor ve maliyet hesaplanıyor...")

def clean_name(x):
    return str(x).replace("\u00a0", " ").strip()

df_noktalar = pd.read_csv(NOKTALAR_FILE, encoding="utf-8-sig")
df_noktalar["Ad"] = df_noktalar["Ad"].apply(clean_name)
df_noktalar["Enlem"]  = pd.to_numeric(df_noktalar["Enlem"].astype(str).str.replace(",", "."), errors="coerce")
df_noktalar["Boylam"] = pd.to_numeric(df_noktalar["Boylam"].astype(str).str.replace(",", "."), errors="coerce")

if df_noktalar[["Enlem", "Boylam"]].isna().any().any():
    print("Uyarı: Bazı koordinatlar sayıya dönüştürülemedi!")

noktalar_dict = df_noktalar.set_index("Ad")[["Enlem", "Boylam"]].to_dict(orient="index")

def calculate_route_distance(route_points):
    total = 0.0
    for i in range(len(route_points) - 1):
        p1, p2 = route_points[i], route_points[i + 1]
        a, b = noktalar_dict.get(p1), noktalar_dict.get(p2)
        if not a or not b:
            raise KeyError(f"Nokta bulunamadı: {p1 if not a else p2}")
        total += geodesic((a["Enlem"], a["Boylam"]), (b["Enlem"], b["Boylam"])).km
    return total

def route_cost(distance_km, num_depots):
    consumption = FUEL_CONSUMPTION_BY_NUM.get(num_depots, 40)
    return (distance_km * (consumption / 100.0)) * FUEL_PRICE_PER_LITRE

def get_capacity(num_depots):
    return CAPACITY_BY_NUM_DEPOTS.get(num_depots, None)

analysis_rows = []
for rid in range(1, TARGET + 1):
    points = all_route_points[rid]
    n_depots = sum(1 for p in points if "DEPO" in str(p).upper())
    try:
        dist_km  = calculate_route_distance(points)
        cost_eur = route_cost(dist_km, n_depots)
        cap      = get_capacity(n_depots)
        err      = ""
    except Exception as e:
        dist_km = cost_eur = cap = None
        err = str(e)

    analysis_rows.append({
        "route_name"     : f"Rota {rid}",
        "hub_point"      : points[0] if points else "",
        "depots"         : " - ".join(p for p in points if "DEPO" in str(p).upper()),
        "num_depots"     : n_depots,
        "distance_km"    : round(dist_km, 4)  if dist_km  is not None else None,
        "route_cost_eur" : round(cost_eur, 4) if cost_eur is not None else None,
        "route_capacity" : cap,
        "error"          : err
    })

df_analysis = pd.DataFrame(analysis_rows)
errors = df_analysis[df_analysis["error"] != ""]
if not errors.empty:
    print(f"Uyarı: {len(errors)} rotada hata var (ilk 3): {errors['error'].head(3).tolist()}")
print("Maliyet hesabı tamamlandı.")

# =========================
# 6) Excel — all sheets in one workbook
# =========================
print("Excel oluşturuluyor...")

# --- Style constants ---
HEADER_FILL  = PatternFill("solid", start_color="2E4057")
SUBHEAD_FILL = PatternFill("solid", start_color="048A81")
ALT_FILL     = PatternFill("solid", start_color="EAF4F4")
WHITE_FILL   = PatternFill("solid", start_color="FFFFFF")
ERR_FILL     = PatternFill("solid", start_color="FFE5E5")

_PALETTE   = ["F4A261","2A9D8F","E76F51","264653","E9C46A",
               "A8DADC","457B9D","E63946","2B9348","9B2226"]
HUB_COLORS = {h: _PALETTE[(h - 1) % len(_PALETTE)] for h in ALL_HUBS}

thin   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_font(size=11, bold=True, color="FFFFFF"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="1A1A2E"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def style_header_row(ws, row, cols, fill=SUBHEAD_FILL):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = hdr_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def style_data_cell(ws, row, col, value, alt=False, bold=False, align="center", err=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = ERR_FILL if err else (ALT_FILL if alt else WHITE_FILL)
    cell.font = body_font(bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = BORDER
    return cell

def add_title_banner(ws, text, merge_range):
    ws.merge_cells(merge_range)
    cell = ws[merge_range.split(":")[0]]
    cell.value = text
    cell.font  = hdr_font(size=14)
    cell.fill  = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")

def freeze_and_size(ws, pane="A3", r1=30, r2=20):
    ws.freeze_panes = pane
    ws.row_dimensions[1].height = r1
    ws.row_dimensions[2].height = r2

wb = Workbook()

# ----------------------------------------------------------
# Sheet 1 — Route Analysis (distance, cost, capacity)
# ----------------------------------------------------------
ws_an = wb.active
ws_an.title = "Route Analysis"
freeze_and_size(ws_an)

add_title_banner(ws_an, f"{TARGET} Route Analysis — Distance, Cost & Capacity", "A1:H1")

hdrs_an = ["Route", "Hub", "Depots Visited", "No. of Depots",
           "Distance (km)", "Cost (EUR)", "Capacity", "Error"]
for col, h in enumerate(hdrs_an, 1):
    ws_an.cell(row=2, column=col, value=h)
style_header_row(ws_an, 2, len(hdrs_an))

for i, rec in enumerate(analysis_rows):
    row = i + 3
    alt = (i % 2 == 1)
    has_err = bool(rec["error"])
    hub_num = route_to_am_out[i + 1]

    style_data_cell(ws_an, row, 1, rec["route_name"],     alt, bold=True, err=has_err)
    hub_cell = style_data_cell(ws_an, row, 2, rec["hub_point"], alt, err=has_err)
    if not has_err:
        hub_cell.fill = PatternFill("solid", start_color=HUB_COLORS[hub_num])
        hub_cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    style_data_cell(ws_an, row, 3, rec["depots"],          alt, align="left", err=has_err)
    style_data_cell(ws_an, row, 4, rec["num_depots"],      alt, err=has_err)
    style_data_cell(ws_an, row, 5, rec["distance_km"],     alt, err=has_err)
    style_data_cell(ws_an, row, 6, rec["route_cost_eur"],  alt, err=has_err)
    style_data_cell(ws_an, row, 7, rec["route_capacity"],  alt, err=has_err)
    style_data_cell(ws_an, row, 8, rec["error"],           alt, align="left", err=has_err)

    # Number formats
    if rec["distance_km"] is not None:
        ws_an.cell(row=row, column=5).number_format = '#,##0.00'
    if rec["route_cost_eur"] is not None:
        ws_an.cell(row=row, column=6).number_format = '#,##0.00'

col_widths_an = [12, 22, 55, 14, 15, 13, 12, 30]
for i, w in enumerate(col_widths_an, 1):
    ws_an.column_dimensions[get_column_letter(i)].width = w

# ----------------------------------------------------------
# Sheet 2 — Raw Data
# ----------------------------------------------------------
ws1 = wb.create_sheet("Raw Data")
freeze_and_size(ws1)

add_title_banner(ws1, f"{TARGET} Route Dataset — Raw Data", "A1:I1")

headers1 = ["Route No", "Starting Hub", "No. of Depots",
            "Depot 1", "Depot 2", "Depot 3", "Depot 4", "Depot 5", "Depot 6"]
for col, h in enumerate(headers1, 1):
    ws1.cell(row=2, column=col, value=h)
style_header_row(ws1, 2, len(headers1))

for i, rid in enumerate(range(1, TARGET + 1)):
    row = i + 3; alt = (i % 2 == 1)
    hub    = route_to_am_out[rid]
    depots = route_to_depots_out[rid]
    style_data_cell(ws1, row, 1, rid, alt, bold=True)
    hub_cell = style_data_cell(ws1, row, 2, hub, alt)
    hub_cell.fill = PatternFill("solid", start_color=HUB_COLORS[hub])
    hub_cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    style_data_cell(ws1, row, 3, len(depots), alt)
    for j, d in enumerate(depots):
        style_data_cell(ws1, row, 4 + j, d, alt)
    for j in range(len(depots), 6):
        style_data_cell(ws1, row, 4 + j, "", alt)

for i, w in enumerate([10, 14, 13, 9, 9, 9, 9, 9, 9], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ----------------------------------------------------------
# Sheet 3 — Python Style
# ----------------------------------------------------------
ws2 = wb.create_sheet("Python Style")
freeze_and_size(ws2)
add_title_banner(ws2, "Python Style Notation", "A1:C1")

for col, h in enumerate(["Route No", "route_no: hub", "route_no: [depots]"], 1):
    ws2.cell(row=2, column=col, value=h)
style_header_row(ws2, 2, 3)

for i, rid in enumerate(range(1, TARGET + 1)):
    row = i + 3; alt = (i % 2 == 1)
    hub    = route_to_am_out[rid]
    depots = route_to_depots_out[rid]
    style_data_cell(ws2, row, 1, rid,               alt, bold=True)
    style_data_cell(ws2, row, 2, f"{rid}: {hub}",   alt, align="left")
    style_data_cell(ws2, row, 3, f"{rid}: {depots}", alt, align="left")

ws2.column_dimensions["A"].width = 10
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 55

# ----------------------------------------------------------
# Sheet 4 — OPL Style
# ----------------------------------------------------------
ws3 = wb.create_sheet("OPL Style")
freeze_and_size(ws3)
add_title_banner(ws3, "OPL Style Notation", "A1:B1")

for col, h in enumerate(["Route No", "<route_no, hub, {depots}>"], 1):
    ws3.cell(row=2, column=col, value=h)
style_header_row(ws3, 2, 2)

for i, rid in enumerate(range(1, TARGET + 1)):
    row = i + 3; alt = (i % 2 == 1)
    hub       = route_to_am_out[rid]
    depots    = route_to_depots_out[rid]
    depot_set = "{" + ",".join(str(d) for d in depots) + "}"
    style_data_cell(ws3, row, 1, rid,                        alt, bold=True)
    style_data_cell(ws3, row, 2, f"<{rid},{hub},{depot_set}>", alt, align="left")

ws3.column_dimensions["A"].width = 10
ws3.column_dimensions["B"].width = 55

# ----------------------------------------------------------
# Sheet 5 — Combined View
# ----------------------------------------------------------
ws4 = wb.create_sheet("Combined View")
freeze_and_size(ws4)
add_title_banner(ws4, "All Notation Styles — Combined View", "A1:D1")

for col, h in enumerate(["Route No", "route_no: hub",
                          "route_no: [depots]", "<route_no, hub, {depots}>"], 1):
    ws4.cell(row=2, column=col, value=h)
style_header_row(ws4, 2, 4)

for i, rid in enumerate(range(1, TARGET + 1)):
    row = i + 3; alt = (i % 2 == 1)
    hub       = route_to_am_out[rid]
    depots    = route_to_depots_out[rid]
    depot_set = "{" + ",".join(str(d) for d in depots) + "}"
    style_data_cell(ws4, row, 1, rid,                          alt, bold=True)
    style_data_cell(ws4, row, 2, f"{rid}: {hub}",             alt, align="left")
    style_data_cell(ws4, row, 3, f"{rid}: {depots}",           alt, align="left")
    style_data_cell(ws4, row, 4, f"<{rid},{hub},{depot_set}>", alt, align="left")

ws4.column_dimensions["A"].width = 10
ws4.column_dimensions["B"].width = 18
ws4.column_dimensions["C"].width = 50
ws4.column_dimensions["D"].width = 55

# ----------------------------------------------------------
# Save
# ----------------------------------------------------------
excel_out = f"routes_{TARGET}_full.xlsx"
wb.save(excel_out)
print(f"✅ Excel yazıldı: {excel_out}")
print(df_analysis[["route_name", "distance_km", "route_cost_eur", "route_capacity"]].head())