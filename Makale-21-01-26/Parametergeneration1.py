"""
VRP Instance Generator + Excel Export
======================================
Adjust the parameters in the CONFIG section below, then run:
    python vrp_generator.py
"""

import random
import math
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ════════════════════════════════════════════════════════════════
#  CONFIG  — change these values as needed
# ════════════════════════════════════════════════════════════════
NUM_SUPPLIERS = 150   # F  (any positive integer)
NUM_DEPOTS    = 40    # D  (must be divisible by 5)
NUM_PERIODS   = 6     # T
SEED          = 42
SCALE_FACTOR  = 1.0   # multiply base demands by this factor
OUTPUT_FILE   = "vrp_instance.xlsx"

# ════════════════════════════════════════════════════════════════
#  CONSTANTS
# ════════════════════════════════════════════════════════════════
VEHICLE_CAPACITY = 245
NUM_SCENARIOS    = 3
SCENARIO_RANGES  = {1: (30, 40), 2: (60, 70), 3: (120, 130)}
SCENARIO_PROBS   = [0.3, 0.5, 0.2]

PATTERN_NAMES = ["decreasing", "increasing", "stationary", "cyclic", "seasonal"]

# Base t=1 demand values for the first 30 depots (from the original data)
BASE_DEMAND_T1 = {
    1: 70,  2: 60,  3: 75,  4: 80,  5: 90,  6: 75,
    7: 30,  8: 18,  9: 40, 10: 80, 11: 22, 12: 20,
   13: 28, 14: 96, 15: 70, 16: 101,17: 50, 18: 90,
   19: 72, 20: 79, 21: 74, 22: 65, 23: 18, 24: 24,
   25: 20, 26: 15, 27: 72, 28: 86, 29: 97, 30: 75,
}

BASE_MEANS = {
    "decreasing": 75.0,
    "increasing": 35.0,
    "stationary": 72.5,
    "cyclic":     60.0,
    "seasonal":   60.8,
}

# ════════════════════════════════════════════════════════════════
#  DEMAND PATTERN FUNCTIONS
# ════════════════════════════════════════════════════════════════

def demand_decreasing(base, t):
    return max(1, round(base * (1 - 0.10) ** (t - 1)))

def demand_increasing(base, t):
    return max(1, round(base * (1 + 0.10) ** (t - 1)))

def demand_stationary(base, t):
    return max(1, round(base * (1 + random.uniform(-0.10, 0.10))))

def demand_cyclic(base, t, T):
    return max(1, round(base * (1 + 0.50 * math.sin(2 * math.pi * (t - 1) / T))))

def demand_seasonal(base, t):
    return max(1, round(base + base * 0.10 * (t - 1)))

PATTERN_FUNCS = {
    "decreasing": demand_decreasing,
    "increasing": demand_increasing,
    "stationary": demand_stationary,
    "cyclic":     demand_cyclic,
    "seasonal":   demand_seasonal,
}

# ════════════════════════════════════════════════════════════════
#  DATA GENERATION
# ════════════════════════════════════════════════════════════════

def get_depot_pattern(d, D):
    """
    Splits D depots into 5 equal groups (D must be divisible by 5):
      Group 1 → decreasing  | Group 2 → increasing | Group 3 → stationary
      Group 4 → cyclic      | Group 5 → seasonal
    """
    assert D % 5 == 0, f"NUM_DEPOTS ({D}) must be divisible by 5."
    group_size = D // 5
    group_idx  = min((d - 1) // group_size, 4)
    return PATTERN_NAMES[group_idx]


def generate_supplier_capacity(F, S, T):
    """capacity[f][s][t] — uniform random within scenario range."""
    cap = {}
    for f in range(1, F + 1):
        cap[f] = {}
        for s in range(1, S + 1):
            lo, hi = SCENARIO_RANGES[s]
            cap[f][s] = {t: random.randint(lo, hi) for t in range(1, T + 1)}
    return cap


def generate_depot_demand(D, T, scale):
    """demand[d][t] — based on pattern group and scaled base value."""
    demand = {}
    for d in range(1, D + 1):
        pattern = get_depot_pattern(d, D)
        func    = PATTERN_FUNCS[pattern]

        if d <= 30:
            base = BASE_DEMAND_T1.get(d, BASE_MEANS[pattern]) * scale
        else:
            base = random.uniform(
                BASE_MEANS[pattern] * 0.7 * scale,
                BASE_MEANS[pattern] * 1.3 * scale,
            )

        demand[d] = {}
        for t in range(1, T + 1):
            if pattern in ("cyclic", "seasonal"):
                demand[d][t] = func(base, t, T) if pattern == "cyclic" else func(base, t)
            else:
                demand[d][t] = func(base, t)
    return demand


def assign_vehicles(F, demand, D, T):
    """
    Assign one vehicle per 4 suppliers (base-case ratio = 25/100).
    Auto-add vehicles until total fleet capacity >= max per-period demand.
    Returns vehicle_supplier[k] = f.
    """
    K = max(1, round(F * 0.25))
    suppliers_with_vehicle = random.sample(range(1, F + 1), min(K, F))
    vehicle_supplier = {k: f for k, f in enumerate(suppliers_with_vehicle, 1)}

    # ensure vehicle capacity covers worst-case period demand
    max_period_demand = max(
        sum(demand[d][t] for d in range(1, D + 1))
        for t in range(1, T + 1)
    )
    while len(vehicle_supplier) * VEHICLE_CAPACITY < max_period_demand:
        k_new = len(vehicle_supplier) + 1
        vehicle_supplier[k_new] = random.randint(1, F)

    return vehicle_supplier


def check_feasibility(capacity, demand, vehicle_supplier, F, D, S, T):
    """Per-period feasibility check for each scenario."""
    K = len(vehicle_supplier)
    total_veh_cap = K * VEHICLE_CAPACITY
    report = {}
    for s in range(1, S + 1):
        per_period = {}
        all_supply_ok = True
        all_vehicle_ok = True
        for t in range(1, T + 1):
            sup_t = sum(capacity[f][s][t] for f in range(1, F + 1))
            dem_t = sum(demand[d][t] for d in range(1, D + 1))
            s_ok  = sup_t >= dem_t
            v_ok  = total_veh_cap >= dem_t
            if not s_ok:  all_supply_ok  = False
            if not v_ok:  all_vehicle_ok = False
            per_period[t] = {"supply": sup_t, "demand": dem_t,
                             "vehicle_cap": total_veh_cap,
                             "supply_ok": s_ok, "vehicle_ok": v_ok}
        report[s] = {"supply_feasible": all_supply_ok,
                     "vehicle_feasible": all_vehicle_ok,
                     "per_period": per_period}
    return report

# ════════════════════════════════════════════════════════════════
#  EXCEL STYLING HELPERS
# ════════════════════════════════════════════════════════════════

C_DARK   = "1F3864"
C_MID    = "2E75B6"
C_LIGHT  = "BDD7EE"
C_S1     = "D9EAD3"
C_S2     = "FFF2CC"
C_S3     = "FCE4D6"
C_TOTAL  = "D6DCE4"
C_ALT    = "F2F2F2"
C_WHITE  = "FFFFFF"
PATTERN_COLORS = {
    "decreasing": "FCE4D6", "increasing": "E2EFDA",
    "stationary": "FFF2CC", "cyclic":     "DAEEF3", "seasonal": "F4CCFF",
}

def ft(bold=False, color="000000", sz=10):
    return Font(name="Arial", bold=bold, color=color, size=sz)

def fill(c):
    return PatternFill("solid", fgColor=c)

def border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="center"):
    return Alignment(horizontal=h, vertical="center", wrap_text=True)

def hcell(ws, r, c, val, bg=C_DARK, fc="FFFFFF", bold=True, sz=10):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = ft(bold=bold, color=fc, sz=sz)
    cell.fill = fill(bg); cell.border = border(); cell.alignment = align()
    return cell

def dcell(ws, r, c, val, bg=C_WHITE, bold=False):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = ft(bold=bold); cell.fill = fill(bg)
    cell.border = border(); cell.alignment = align()
    return cell

def merge_title(ws, r, c1, c2, val, bg=C_DARK):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(row=r, column=c1, value=val)
    cell.font = ft(bold=True, color="FFFFFF", sz=11)
    cell.fill = fill(bg); cell.border = border(); cell.alignment = align()

def cw(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

# ════════════════════════════════════════════════════════════════
#  SHEET 1 — Supplier Capacity  c(f, s, t)
# ════════════════════════════════════════════════════════════════

def write_capacity_sheet(ws, capacity, F, S, T):
    ws.title = "Supplier Capacity"
    S_COLORS = [C_S1, C_S2, C_S3]
    S_LABELS = ["Scenario 1  [30–40]", "Scenario 2  [60–70]", "Scenario 3  [120–130]"]

    total_cols = 1 + S * (T + 1)
    merge_title(ws, 1, 1, total_cols, "Supplier Capacity  –  c(f, s, t)")
    ws.row_dimensions[1].height = 22

    # Row 2: scenario headers
    col = 2
    for s in range(1, S + 1):
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + T)
        hcell(ws, 2, col, S_LABELS[s-1], bg=S_COLORS[s-1], fc="000000")
        col += T + 1
    ws.row_dimensions[2].height = 18

    # Row 3: period sub-headers
    hcell(ws, 3, 1, "Supplier (f)", bg=C_LIGHT, fc="000000")
    col = 2
    for s in range(1, S + 1):
        for t in range(1, T + 1):
            hcell(ws, 3, col, f"t = {t}", bg=S_COLORS[s-1], fc="000000"); col += 1
        hcell(ws, 3, col, "Total", bg=C_TOTAL, fc="000000");               col += 1
    ws.row_dimensions[3].height = 16

    # Data rows
    for f in range(1, F + 1):
        r  = 3 + f
        bg = C_WHITE if f % 2 == 0 else C_ALT
        hcell(ws, r, 1, f"f = {f}", bg=C_LIGHT, fc="000000")
        col = 2
        for s in range(1, S + 1):
            t_cols = []
            for t in range(1, T + 1):
                dcell(ws, r, col, capacity[f][s][t], bg=bg); t_cols.append(get_column_letter(col)); col += 1
            dcell(ws, r, col, f"=SUM({t_cols[0]}{r}:{t_cols[-1]}{r})", bg=C_TOTAL, bold=True); col += 1

    # Total row
    tr = 3 + F + 1
    hcell(ws, tr, 1, "TOTAL", bg=C_DARK)
    col = 2
    for _ in range(S * (T + 1)):
        cl = get_column_letter(col)
        dcell(ws, tr, col, f"=SUM({cl}4:{cl}{tr-1})", bg=C_TOTAL, bold=True); col += 1

    cw(ws, 1, 16)
    for c in range(2, total_cols + 1): cw(ws, c, 10)
    ws.freeze_panes = "B4"

# ════════════════════════════════════════════════════════════════
#  SHEET 2 — Depot Demand  d(d, t)
# ════════════════════════════════════════════════════════════════

def write_demand_sheet(ws, demand, D, T):
    ws.title = "Depot Demand"
    total_cols = 3 + T + 1
    merge_title(ws, 1, 1, total_cols, "Depot Demand  –  d(d, t)")
    ws.row_dimensions[1].height = 22

    for c, h in enumerate(["Depot (d)", "Pattern", "Group",
                            *[f"t = {t}" for t in range(1, T + 1)], "Total"], 1):
        hcell(ws, 2, c, h, bg=C_LIGHT, fc="000000")
    ws.row_dimensions[2].height = 16

    group_size = D // 5
    group_labels = ["G1 – Decreasing", "G2 – Increasing", "G3 – Stationary",
                    "G4 – Cyclic",     "G5 – Seasonal"]

    for d in range(1, D + 1):
        r       = 2 + d
        pattern = get_depot_pattern(d, D)
        g_idx   = min((d - 1) // group_size, 4)
        bg      = PATTERN_COLORS[pattern]

        hcell(ws, r, 1, f"d = {d}",         bg=bg, fc="000000", bold=False)
        dcell(ws, r, 2, pattern.capitalize(), bg=bg)
        dcell(ws, r, 3, group_labels[g_idx],  bg=bg)

        t0, tN = 4, 3 + T
        for t in range(1, T + 1):
            dcell(ws, r, t0 + t - 1, demand[d][t], bg=bg)

        s_col = get_column_letter(t0); e_col = get_column_letter(tN)
        dcell(ws, r, tN + 1, f"=SUM({s_col}{r}:{e_col}{r})", bg=C_TOTAL, bold=True)

    # Total row
    tr = 2 + D + 1
    hcell(ws, tr, 1, "TOTAL")
    for c in range(2, 3): dcell(ws, tr, c, "", bg=C_TOTAL)
    for c in range(4, tN + 2):
        cl = get_column_letter(c)
        dcell(ws, tr, c, f"=SUM({cl}3:{cl}{tr-1})", bg=C_TOTAL, bold=True)

    # Legend
    lr = tr + 2
    ws.cell(row=lr, column=1, value="Legend").font = ft(bold=True)
    for i, pat in enumerate(PATTERN_NAMES):
        cell = ws.cell(row=lr + 1 + i, column=1,
                       value=f"{group_labels[i]}  →  {pat.capitalize()}")
        cell.fill = fill(PATTERN_COLORS[pat]); cell.font = ft(); cell.border = border()
        cell.alignment = align(h="left")

    cw(ws, 1, 14); cw(ws, 2, 16); cw(ws, 3, 20)
    for c in range(4, total_cols + 1): cw(ws, c, 10)
    ws.freeze_panes = "D3"

# ════════════════════════════════════════════════════════════════
#  SHEET 3 — Vehicles & Balance Check
# ════════════════════════════════════════════════════════════════

def write_vehicle_sheet(ws, vehicle_supplier, feasibility, S, T):
    ws.title = "Vehicles & Balance"
    K = len(vehicle_supplier)

    # Section A: vehicle assignment
    merge_title(ws, 1, 1, 4,
                f"Vehicle Assignment  |  K = {K}  |  Capacity per vehicle = {VEHICLE_CAPACITY}")
    ws.row_dimensions[1].height = 22
    for c, h in enumerate(["Vehicle (k)", "Supplier (f)", "Capacity", "Cumulative Cap"], 1):
        hcell(ws, 2, c, h, bg=C_LIGHT, fc="000000")

    for i, (k, f) in enumerate(sorted(vehicle_supplier.items())):
        r  = 3 + i
        bg = "E2EFDA" if i % 2 == 0 else C_WHITE
        dcell(ws, r, 1, f"k = {k}", bg=bg)
        dcell(ws, r, 2, f"f = {f}", bg=bg)
        dcell(ws, r, 3, VEHICLE_CAPACITY, bg=bg)
        dcell(ws, r, 4, f"=SUM(C3:C{r})", bg=bg, bold=False)

    tr_v = 3 + K
    hcell(ws, tr_v, 1, "TOTAL")
    dcell(ws, tr_v, 2, "", bg=C_TOTAL)
    dcell(ws, tr_v, 3, f"=SUM(C3:C{tr_v-1})", bg=C_TOTAL, bold=True)
    dcell(ws, tr_v, 4, "", bg=C_TOTAL)
    for c in range(1, 5): cw(ws, c, 18)

    # Section B: balance check per scenario
    S_COLORS = [C_S1, C_S2, C_S3]
    S_LABELS = ["Scenario 1  [30–40]", "Scenario 2  [60–70]", "Scenario 3  [120–130]"]
    bc = 6   # starting column for balance tables

    for s in range(1, S + 1):
        merge_title(ws, 1, bc, bc + 5, S_LABELS[s-1], bg=C_MID)
        ws.cell(row=1, column=bc).fill  = fill(S_COLORS[s-1])
        ws.cell(row=1, column=bc).font  = ft(bold=True, color="000000", sz=10)
        for c2, h in enumerate(["Period (t)", "Supply (Σf)", "Demand (Σd)",
                                 "Vehicle Cap", "Supply OK?", "Vehicle OK?"], bc):
            hcell(ws, 2, c2, h, bg=S_COLORS[s-1], fc="000000")

        for t in range(1, T + 1):
            r   = 2 + t
            bg  = C_WHITE if t % 2 == 0 else C_ALT
            pp  = feasibility[s]["per_period"][t]
            sup_ok = pp["supply_ok"]; veh_ok = pp["vehicle_ok"]

            dcell(ws, r, bc,     f"t = {t}",         bg=bg)
            dcell(ws, r, bc + 1, pp["supply"],        bg=bg)
            dcell(ws, r, bc + 2, pp["demand"],        bg=bg)
            dcell(ws, r, bc + 3, pp["vehicle_cap"],   bg=bg)

            for offset, ok in [(4, sup_ok), (5, veh_ok)]:
                c_letter_sup = get_column_letter(bc + 1)
                c_letter_dem = get_column_letter(bc + 2)
                c_letter_veh = get_column_letter(bc + 3)
                if offset == 4:
                    formula = f'=IF({c_letter_sup}{r}>={c_letter_dem}{r},"✓ OK","✗ FAIL")'
                else:
                    formula = f'=IF({c_letter_veh}{r}>={c_letter_dem}{r},"✓ OK","✗ FAIL")'
                cell = ws.cell(row=r, column=bc + offset, value=formula)
                cell.font      = ft(bold=True, color="006100" if ok else "9C0006")
                cell.fill      = fill("C6EFCE" if ok else "FFC7CE")
                cell.border    = border(); cell.alignment = align()

        # Total row
        tr_b = 2 + T + 1
        hcell(ws, tr_b, bc, "TOTAL")
        for offset in range(1, 4):
            cl = get_column_letter(bc + offset)
            dcell(ws, tr_b, bc + offset, f"=SUM({cl}3:{cl}{tr_b-1})", bg=C_TOTAL, bold=True)
        for offset in range(4, 6):
            dcell(ws, tr_b, bc + offset, "", bg=C_TOTAL)

        for j in range(6): cw(ws, bc + j, 16)
        bc += 7   # gap between scenario blocks

    ws.freeze_panes = "A3"

# ════════════════════════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════════════════════════


# ================================================================
#  SHEET 4 - Upsilon  (flat table: f, s, t, supplier_capacity)
# ================================================================

def write_upsilon_sheet(ws, capacity, F, S, T):
    ws.title = "Upsilon"

    merge_title(ws, 1, 1, 4, "Upsilon  -  Supplier Capacity  (flat table)")
    ws.row_dimensions[1].height = 22

    for c, h in enumerate(["f  (Supplier)", "s  (Scenario)", "t  (Period)", "Capacity"], 1):
        hcell(ws, 2, c, h, bg=C_LIGHT, fc="000000")
    ws.row_dimensions[2].height = 16

    S_COLORS = [C_S1, C_S2, C_S3]
    r = 3
    for f in range(1, F + 1):
        for s in range(1, S + 1):
            for t in range(1, T + 1):
                bg = S_COLORS[s - 1]
                dcell(ws, r, 1, f,                  bg=bg)
                dcell(ws, r, 2, s,                  bg=bg)
                dcell(ws, r, 3, t,                  bg=bg)
                dcell(ws, r, 4, capacity[f][s][t],  bg=bg)
                r += 1

    # Total row
    dcell(ws, r, 1, "TOTAL",            bg=C_TOTAL, bold=True)
    dcell(ws, r, 2, "",                 bg=C_TOTAL)
    dcell(ws, r, 3, "",                 bg=C_TOTAL)
    dcell(ws, r, 4, f"=SUM(D3:D{r-1})", bg=C_TOTAL, bold=True)

    cw(ws, 1, 16); cw(ws, 2, 16); cw(ws, 3, 14); cw(ws, 4, 18)
    ws.freeze_panes = "A3"

def main():
    assert NUM_DEPOTS % 5 == 0, f"NUM_DEPOTS ({NUM_DEPOTS}) must be divisible by 5."

    random.seed(SEED)
    F, D, S, T = NUM_SUPPLIERS, NUM_DEPOTS, NUM_SCENARIOS, NUM_PERIODS

    print(f"Generating instance: F={F}, D={D}, T={T}, S={S}, scale={SCALE_FACTOR}")

    capacity        = generate_supplier_capacity(F, S, T)
    demand          = generate_depot_demand(D, T, SCALE_FACTOR)
    vehicle_supplier = assign_vehicles(F, demand, D, T)
    feasibility     = check_feasibility(capacity, demand, vehicle_supplier, F, D, S, T)

    K = len(vehicle_supplier)
    print(f"Vehicles assigned: K={K}  |  Total fleet capacity: {K * VEHICLE_CAPACITY}")
    for s in range(1, S + 1):
        r = feasibility[s]
        print(f"  Scenario {s}: supply={'✓' if r['supply_feasible'] else '✗'}  "
              f"vehicle={'✓' if r['vehicle_feasible'] else '✗'}")

    # Write Excel
    wb = Workbook()
    wb.remove(wb.active)

    write_capacity_sheet(wb.create_sheet("Supplier Capacity"), capacity, F, S, T)
    write_demand_sheet(wb.create_sheet("Depot Demand"),        demand,   D, T)
    write_vehicle_sheet(wb.create_sheet("Vehicles & Balance"), vehicle_supplier, feasibility, S, T)
    write_upsilon_sheet(wb.create_sheet("Upsilon"),            capacity, F, S, T)

    wb.save(OUTPUT_FILE)
    print(f"\nSaved → {OUTPUT_FILE}")


if __name__ == "__main__":
    main()