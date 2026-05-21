"""
=============================================================================
Threshold-Based Collaborative Selling Problem
REVISED MODEL — All-unit threshold pricing with collaboration incentives
=============================================================================
"""

import math
from docplex.mp.model import Model


# =============================================================================
# 1. INSTANCE DATA
# =============================================================================

producers = list(range(1, 11))
levels = [0, 1, 2]
G = list(range(len(producers)))

district = {
    1: "Torbalı, İzmir", 2: "Selçuk, İzmir", 3: "Urla, İzmir",
    4: "Foça, İzmir", 5: "Ödemiş, İzmir", 6: "Nazilli, Aydın",
    7: "Söke, Aydın", 8: "Akhisar, Manisa", 9: "Kırkağaç, Manisa",
    10: "Milas, Muğla",
}

lat = {
    1: 38.163, 2: 37.956, 3: 38.321, 4: 38.673, 5: 38.226,
    6: 37.913, 7: 37.748, 8: 38.919, 9: 39.101, 10: 37.318
}

lon = {
    1: 27.359, 2: 27.369, 3: 26.765, 4: 26.755, 5: 27.971,
    6: 28.317, 7: 27.413, 8: 27.837, 9: 27.668, 10: 27.787
}

q = {1: 12, 2: 7, 3: 15, 4: 10, 5: 5, 6: 18, 7: 6, 8: 14, 9: 10, 10: 9}
#q = {1: 12, 2: 7, 3: 11, 4: 8, 5: 5, 6: 13, 7: 6, 8: 9, 9: 10, 10: 9}



def haversine(la1, lo1, la2, lo2):
    R = 6371.0
    phi1, phi2 = math.radians(la1), math.radians(la2)
    a = (
        math.sin(math.radians(la2 - la1) / 2) ** 2
        + math.cos(phi1) * math.cos(phi2)
        * math.sin(math.radians(lo2 - lo1) / 2) ** 2
    )
    return 2.0 * R * math.asin(math.sqrt(a))


d = {
    (i, j): haversine(lat[i], lon[i], lat[j], lon[j])
    for i in producers for j in producers if i != j
}

D_max = 75.0

infeasible_pairs = [
    (i, j)
    for i in producers for j in producers
    if i < j and d[(i, j)] > D_max
]


# =============================================================================
# 2. THRESHOLD-BASED REVENUE PARAMETERS
# =============================================================================

A = {
    0: 0,     # base price level
    1: 25,    # Medium quantity price level
    2: 30,    # High quantity price level
}

p = {
    0: 130,
    1: 175,
    2: 220,
}

price_level = {
    0: "Base price level",
    1: "Medium quantity price level",
    2: "High quantity price level",
}

M = sum(q.values())


# =============================================================================
# 3. MODEL
# =============================================================================

mdl = Model(name="Threshold_Based_Collab_Selling_N10_L3")
#mdl.parameters.mip.tolerances.mipgap = 1e-6
#mdl.parameters.timelimit = 3600

x = {
    (i, g): mdl.binary_var(name=f"x_{i}_{g}")
    for i in producers for g in G
}

s = {
    i: mdl.binary_var(name=f"s_{i}")
    for i in producers
}

y = {
    g: mdl.binary_var(name=f"y_{g}")
    for g in G
}

Q = {
    g: mdl.continuous_var(lb=0, name=f"Q_{g}")
    for g in G
}

zG = {
    (g, l): mdl.binary_var(name=f"zG_{g}_{l}")
    for g in G for l in levels
}

zI = {
    (i, l): mdl.binary_var(name=f"zI_{i}_{l}")
    for i in producers for l in levels
}

wG = {
    (g, l): mdl.continuous_var(lb=0, name=f"wG_{g}_{l}")
    for g in G for l in levels
}


# =============================================================================
# 4. OBJECTIVE FUNCTION
# =============================================================================

mdl.maximize(
    mdl.sum(p[l] * wG[(g, l)] for g in G for l in levels)
    + mdl.sum(p[l] * q[i] * zI[(i, l)] for i in producers for l in levels)
)


# =============================================================================
# 5. CONSTRAINTS
# =============================================================================

for i in producers:
    mdl.add_constraint(
        mdl.sum(x[(i, g)] for g in G) + s[i] == 1,
        ctname=f"assignment_i{i}"
    )

for i in producers:
    for g in G:
        mdl.add_constraint(
            x[(i, g)] <= y[g],
            ctname=f"active_assignment_i{i}_g{g}"
        )

for g in G:
    mdl.add_constraint(
        mdl.sum(x[(i, g)] for i in producers) >= 2 * y[g],
        ctname=f"min_group_size_g{g}"
    )

for g in G:
    mdl.add_constraint(
        mdl.sum(x[(i, g)] for i in producers) <= len(producers) * y[g],
        ctname=f"empty_group_inactive_g{g}"
    )

for g in G:
    mdl.add_constraint(
        Q[g] == mdl.sum(q[i] * x[(i, g)] for i in producers),
        ctname=f"group_quantity_g{g}"
    )

for g in G:
    mdl.add_constraint(
        mdl.sum(zG[(g, l)] for l in levels) == y[g],
        ctname=f"one_level_group_g{g}"
    )

for i in producers:
    mdl.add_constraint(
        mdl.sum(zI[(i, l)] for l in levels) == s[i],
        ctname=f"one_level_individual_i{i}"
    )

for g in G:
    for l in levels:
        mdl.add_constraint(
            Q[g] >= A[l] * zG[(g, l)],
            ctname=f"group_threshold_g{g}_l{l}"
        )

for i in producers:
    for l in levels:
        mdl.add_constraint(
            q[i] >= A[l] * zI[(i, l)],
            ctname=f"individual_threshold_i{i}_l{l}"
        )

for g in G:
    for l in levels:
        mdl.add_constraint(
            wG[(g, l)] <= Q[g],
            ctname=f"w_le_Q_g{g}_l{l}"
        )
        mdl.add_constraint(
            wG[(g, l)] <= M * zG[(g, l)],
            ctname=f"w_le_Mz_g{g}_l{l}"
        )
        mdl.add_constraint(
            wG[(g, l)] >= Q[g] - M * (1 - zG[(g, l)]),
            ctname=f"w_ge_Q_minus_M_g{g}_l{l}"
        )

for g in G:
    for i, j in infeasible_pairs:
        mdl.add_constraint(
            x[(i, g)] + x[(j, g)] <= 1,
            ctname=f"geo_i{i}_j{j}_g{g}"
        )


# =============================================================================
# 6. SOLVE
# =============================================================================

print("Solving with IBM CPLEX...")
sol = mdl.solve(log_output=True)


# =============================================================================
# 7. SOLUTION REPORT
# =============================================================================

def selected_individual_level(i, sol):
    for l in levels:
        if sol.get_value(zI[(i, l)]) > 0.5:
            return l
    return None


def selected_group_level(g, sol):
    for l in levels:
        if sol.get_value(zG[(g, l)]) > 0.5:
            return l
    return None


if sol is None:
    print("No solution found.")

else:
    print("\n" + "=" * 70)
    print("SOLUTION REPORT")
    print("=" * 70)

    print(f"Solver status   : {mdl.solve_details.status}")
    print(f"Optimal revenue : ${sol.objective_value:,.2f}")

    ind = [i for i in producers if sol.get_value(s[i]) > 0.5]
    act = [g for g in G if sol.get_value(y[g]) > 0.5]

    print(f"\nINDIVIDUAL SELLERS ({len(ind)})")

    individual_revenue = 0

    for i in ind:
        l = selected_individual_level(i, sol)
        rev_i = p[l] * q[i]
        individual_revenue += rev_i

        print(
            f"Producer {i} | {district[i]} | "
            f"q={q[i]} | level={l} | price={p[l]} | revenue={rev_i}"
        )

    print(f"\nCOLLABORATIVE GROUPS ({len(act)})")

    group_revenue = 0

    for g in act:
        members = [i for i in producers if sol.get_value(x[(i, g)]) > 0.5]
        total_q = sum(q[i] for i in members)
        l = selected_group_level(g, sol)
        rev_g = p[l] * total_q
        group_revenue += rev_g

        print(f"\nGroup {g}")
        print(f"Members: {members}")
        print(f"Districts: {[district[i] for i in members]}")
        print(f"Total quantity: {total_q}")
        print(f"Selected level: {l}")
        print(f"Threshold: {A[l]}")
        print(f"Unit price: {p[l]}")
        print(f"Revenue: {rev_g}")

    print("\nREVENUE SUMMARY")
    print(f"Individual revenue: ${individual_revenue:,.2f}")
    print(f"Group revenue     : ${group_revenue:,.2f}")
    print(f"Total revenue     : ${sol.objective_value:,.2f}")


# =============================================================================
# 8. EXPORT RESULTS TO EXCEL
# =============================================================================

def write_results_to_excel(sol, filepath="threshold_model_results.xlsx"):
    import openpyxl

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "x_ig"
    ws.cell(row=1, column=1, value="x_ig")
    ws.cell(row=2, column=1, value="i \\ g")

    for g in G:
        ws.cell(row=2, column=g + 2, value=g)

    for row_idx, i in enumerate(producers, start=3):
        ws.cell(row=row_idx, column=1, value=i)
        for col_idx, g in enumerate(G, start=2):
            ws.cell(row=row_idx, column=col_idx, value=round(sol.get_value(x[(i, g)])))

    ws2 = wb.create_sheet("s_i")
    ws2.append(["i", "district", "q_i", "s_i", "selected_level", "unit_price", "revenue"])

    for i in producers:
        if sol.get_value(s[i]) > 0.5:
            l = selected_individual_level(i, sol)
            ws2.append([i, district[i], q[i], 1, l, p[l], p[l] * q[i]])
        else:
            ws2.append([i, district[i], q[i], 0, None, None, 0])

    ws3 = wb.create_sheet("y_g")
    ws3.append(["g", "y_g", "Q_g", "selected_level", "unit_price", "revenue", "members"])

    for g in G:
        if sol.get_value(y[g]) > 0.5:
            l = selected_group_level(g, sol)
            members = [i for i in producers if sol.get_value(x[(i, g)]) > 0.5]
            Q_val = sol.get_value(Q[g])
            ws3.append([
                g, 1, Q_val, l, p[l], p[l] * Q_val,
                ", ".join(map(str, members))
            ])
        else:
            ws3.append([g, 0, 0, None, None, 0, ""])

    ws4 = wb.create_sheet("zG_gl")
    ws4.append(["g \\ l"] + levels)

    for g in G:
        ws4.append([g] + [round(sol.get_value(zG[(g, l)])) for l in levels])

    ws5 = wb.create_sheet("zI_il")
    ws5.append(["i \\ l"] + levels)

    for i in producers:
        ws5.append([i] + [round(sol.get_value(zI[(i, l)])) for l in levels])

    ws6 = wb.create_sheet("wG_gl")
    ws6.append(["g \\ l"] + levels)

    for g in G:
        ws6.append([g] + [round(sol.get_value(wG[(g, l)]), 4) for l in levels])

    ws7 = wb.create_sheet("price_levels")
    ws7.append(["level", "price_level_description", "minimum_quantity_A_l", "unit_price_p_l"])

    for l in levels:
        ws7.append([l, price_level[l], A[l], p[l]])

    wb.save(filepath)
    print(f"\nResults saved to: {filepath}")


if sol is not None:
    write_results_to_excel(sol)